"""Tests for CN Fund format validator.

Uses real Excel fixtures to verify sheet and column presence.
"""
import pytest

pytestmark = pytest.mark.pipeline

import openpyxl
from src.validation.source_format_validator import validate_cn_fund_format

@pytest.fixture
def valid_cn_fund_xlsx(tmp_path):
    """Create a valid CN Fund workbook."""
    path = tmp_path / "valid_fund.xlsx"
    wb = openpyxl.Workbook()
    
    # Sheet 1: Holdings (English headers)
    ws1 = wb.active
    ws1.title = "基金持仓汇总"
    ws1.append(["Asset_ID", "Asset_Name", "Quantity", "Market_Price_Unit", "Market_Value_Raw"])
    
    # Sheet 2: Transactions (Chinese headers)
    ws2 = wb.create_sheet("基金交易记录")
    ws2.append(["交易日期", "基金代码", "基金名称", "操作类型", "交易金额"])
    
    wb.save(path)
    return path

@pytest.fixture
def missing_sheet_xlsx(tmp_path):
    """Create a workbook missing one sheet."""
    path = tmp_path / "missing_sheet.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "基金持仓汇总"
    wb.save(path)
    return path

def test_validate_valid_file(valid_cn_fund_xlsx):
    result = validate_cn_fund_format(valid_cn_fund_xlsx)
    assert result.is_valid is True
    assert result.file_type == "cn_fund"
    assert len(result.warnings) == 0

def test_validate_missing_file(tmp_path):
    result = validate_cn_fund_format(tmp_path / "nonexistent.xlsx")
    assert result.is_valid is False
    assert "File not found" in result.warnings[0]

def test_validate_missing_sheet(missing_sheet_xlsx):
    result = validate_cn_fund_format(missing_sheet_xlsx)
    assert result.is_valid is False
    assert "Missing required sheet: 基金交易记录" in result.warnings[0]

def test_validate_missing_columns(tmp_path):
    path = tmp_path / "bad_cols.xlsx"
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "基金持仓汇总"
    ws1.append(["Asset_ID", "Wrong_Col"]) # Missing others
    ws2 = wb.create_sheet("基金交易记录")
    ws2.append(["交易日期", "基金代码"]) # Missing others
    wb.save(path)
    
    result = validate_cn_fund_format(path)
    assert result.is_valid is False
    assert any("Missing holdings columns" in w for w in result.warnings)
    assert any("Missing transaction columns" in w for w in result.warnings)

def test_validate_non_excel_file(tmp_path):
    path = tmp_path / "not_excel.txt"
    path.write_text("hello world")
    
    result = validate_cn_fund_format(path)
    assert result.is_valid is False
    assert "Cannot read Excel file" in result.warnings[0]
