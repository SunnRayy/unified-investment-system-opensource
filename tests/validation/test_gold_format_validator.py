"""Tests for Gold format validator."""
import pytest

pytestmark = pytest.mark.pipeline

import openpyxl
from src.validation.source_format_validator import validate_gold_format

@pytest.fixture
def valid_gold_workbook(tmp_path):
    wb = openpyxl.Workbook()
    ws_h = wb.active
    ws_h.title = "黄金持仓"
    ws_h.append(["资产类别", "标的名称", "持有数量", "单位", "平均成本价",
                  "单价", "当前市值", "未实现盈亏", "交易账户"])
    ws_h.append(["黄金", "纸黄金", 100.0, "克", 500, 600, 60000, 10000, "招行"])
    ws_t = wb.create_sheet("黄金交易记录")
    ws_t.append(["交易日期", "资产类别", "标的名称", "交易类型", "金额",
                  "数量", "价格", "手续费", "交易账户"])
    path = tmp_path / "Gold_transactions.xlsx"
    wb.save(path)
    return path

class TestGoldFormatValidator:
    def test_valid_workbook(self, valid_gold_workbook):
        result = validate_gold_format(valid_gold_workbook)
        assert result.is_valid is True
        assert result.file_type == "gold"

    def test_missing_sheet(self, tmp_path):
        wb = openpyxl.Workbook()
        wb.active.title = "WrongSheet"
        path = tmp_path / "bad.xlsx"
        wb.save(path)
        result = validate_gold_format(path)
        assert result.is_valid is False

    def test_file_not_found(self, tmp_path):
        result = validate_gold_format(tmp_path / "nonexistent.xlsx")
        assert result.is_valid is False
