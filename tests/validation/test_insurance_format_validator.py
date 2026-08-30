"""Tests for Insurance format validator."""
import pytest

pytestmark = pytest.mark.pipeline

import openpyxl
from src.validation.source_format_validator import validate_insurance_format

@pytest.fixture
def valid_insurance_workbook(tmp_path):
    wb = openpyxl.Workbook()
    ws_s = wb.active
    ws_s.title = "保险汇总"
    ws_s.append(["产品名称", "保险公司", "产品类型", "开始日期", "保障期限",
                  "缴费期限", "年保费", "保额", "保障范围", "保单状态"])
    ws_p = wb.create_sheet("保费记录")
    ws_p.append(["日期", "TestPolicy"])
    path = tmp_path / "Insurance_Portfolio.xlsx"
    wb.save(path)
    return path

class TestInsuranceFormatValidator:
    def test_valid(self, valid_insurance_workbook):
        result = validate_insurance_format(valid_insurance_workbook)
        assert result.is_valid is True
        assert result.file_type == "insurance"

    def test_missing_sheet(self, tmp_path):
        wb = openpyxl.Workbook()
        wb.active.title = "Wrong"
        path = tmp_path / "bad.xlsx"
        wb.save(path)
        assert validate_insurance_format(path).is_valid is False

    def test_file_not_found(self, tmp_path):
        assert validate_insurance_format(tmp_path / "nope.xlsx").is_valid is False
