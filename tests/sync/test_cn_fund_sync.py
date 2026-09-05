"""Tests for CN Fund sync module.

Tests the sync_cn_fund(config) interface.
"""
import pytest

pytestmark = pytest.mark.pipeline

import openpyxl
from src.sync.cn_fund_sync import sync_cn_fund

@pytest.fixture
def sync_config(tmp_path):
    """Create a mock config for sync tests."""
    data_dir = tmp_path / "data"
    data_dir.mkdir()
    
    # Create valid workbook
    wb_path = data_dir / "funding_transactions.xlsx"
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "基金持仓汇总"
    ws1.append(["Asset_ID", "Asset_Name", "Quantity", "Market_Price_Unit", "Market_Value_Raw"])
    ws1.append(["900001", "Fund A", 100, 1.0, 100])
    ws2 = wb.create_sheet("基金交易记录")
    ws2.append(["交易日期", "基金代码", "基金名称", "操作类型", "交易金额"])
    wb.save(wb_path)
    
    return {
        'source_registry': {
            'cn_fund': {
                'enabled': True,
                'data_dir': str(data_dir),
                'file_patterns': {'workbook': 'funding_transactions.xlsx'}
            }
        }
    }

def test_sync_cn_fund_success(sync_config):
    result = sync_cn_fund(sync_config)
    
    assert isinstance(result, dict)
    assert "holdings" in result
    assert "transactions" in result
    assert len(result["holdings"]) == 1
    assert result["holdings"].iloc[0]["asset_id"] == "CN_FUND_900001"

def test_sync_cn_fund_disabled(sync_config):
    sync_config['source_registry']['cn_fund']['enabled'] = False
    result = sync_cn_fund(sync_config)
    
    assert result["holdings"].empty
    assert result["transactions"].empty

def test_sync_cn_fund_file_not_found(sync_config):
    # Change filename to something that doesn't exist
    sync_config['source_registry']['cn_fund']['file_patterns']['workbook'] = "missing.xlsx"
    result = sync_cn_fund(sync_config)
    
    assert result["holdings"].empty
    assert result["transactions"].empty

def test_sync_cn_fund_finance_dir_fallback(tmp_path):
    """Test fallback to finance_dir (iCloud) if data_dir is null.

    Phase 9 (PIS deprecation): The old PIS subsystem path fallback is replaced
    by finance_dir so that all readers share a single iCloud-based fallback.
    """
    finance_dir = tmp_path / "Finance"
    finance_dir.mkdir()

    # Create workbook in finance_dir
    wb_path = finance_dir / "funding_transactions.xlsx"
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "基金持仓汇总"
    ws1.append(["Asset_ID", "Asset_Name", "Quantity", "Market_Price_Unit", "Market_Value_Raw"])
    ws1.append(["900001", "Fund A", 100, 1.0, 100])
    wb.save(wb_path)

    config = {
        'source_registry': {
            'cn_fund': {
                'enabled': True,
                'data_dir': None,
                'file_patterns': {'workbook': 'funding_transactions.xlsx'}
            }
        },
        'finance_dir': str(finance_dir),
    }

    result = sync_cn_fund(config)
    assert len(result["holdings"]) == 1


def test_uses_finance_dir_when_data_dir_null(tmp_path):
    """When data_dir is null, falls back to finance_dir (iCloud), not PIS path.

    The key assertion: we verify the file in finance_dir was actually found and
    read, not that we fell through to the early-return-empty path.  We do this
    by putting a row in the workbook and checking the result is NOT empty — if
    the implementation were returning empty because the path wasn't resolved,
    this assertion would fail.
    """
    finance_dir = tmp_path / "Finance"
    finance_dir.mkdir()

    # Create a minimal funding_transactions.xlsx with the expected sheet and columns.
    # Uses the same processed-format column names as CNFundReader expects.
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "基金持仓汇总"
    ws.append(["Asset_ID", "Asset_Name", "Quantity", "Market_Price_Unit", "Market_Value_Raw"])
    ws.append(["900001", "Fund A", 100, 1.0, 100])  # one data row
    wb.save(str(finance_dir / "funding_transactions.xlsx"))

    config = {
        'source_registry': {
            'cn_fund': {
                'enabled': True,
                'data_dir': None,
                'file_patterns': {'workbook': 'funding_transactions.xlsx'},
            }
        },
        'finance_dir': str(finance_dir),
        # No sources.pis fallback config
        # No subsystems.pis.path
    }

    result = sync_cn_fund(config)
    # Must contain 'holdings' key
    assert 'holdings' in result
    # The workbook in finance_dir was found and read — not an early-exit empty result
    assert not result['holdings'].empty, (
        "Expected finance_dir to be used; got empty holdings which means "
        "the implementation returned early without finding the file."
    )
