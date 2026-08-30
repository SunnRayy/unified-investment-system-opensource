"""Tests for src/sync/reference_export.py — D1 reference-sheet generator.

Safety rules enforced throughout:
- Never write to the real Finance dir or any production workbook.
- All DB access uses an in-memory DuckDB fixture.
- tmp_path (pytest) is the only allowed output location.
"""
from __future__ import annotations

import textwrap
from pathlib import Path
from typing import Any, Dict
from unittest.mock import MagicMock, patch

import pytest

from src.database.connector import DatabaseConnector
from src.database.schema import initialize_schema
from src.sync.reference_export import (
    MONTHLY_SHEET_NAME,
    OUTPUT_FILENAME,
    _assert_safe_output_path,
    _build_rows,
    _compute_monthly_snapshot_rows,
    _get_fx_rate,
    _get_market_price,
    _load_reference_config,
    _merge_monthly_snapshot,
    _query_asset,
    _query_broker_consolidated_share,
    _query_source_sum,
    export_reference_sheet,
)

# ── Fixtures ──────────────────────────────────────────────────────────────────

@pytest.fixture
def mem_db():
    """In-memory DuckDB with schema, pre-seeded with small holdings fixture."""
    conn = DatabaseConnector(":memory:")
    initialize_schema(conn)

    # Seed asset_registry
    conn.execute("""
        INSERT INTO asset_registry (canonical_id, display_name, base_currency, is_active)
        VALUES
            ('RSU_AMZN',       'Amazon RSU',    'USD', TRUE),
            ('RSU_GOOGL',      'Google RSU',    'USD', TRUE),
            ('ALTS_Paper_Gold','Paper Gold',    'CNY', TRUE),
            ('CN_FUND_001',    'CN Fund A',     'CNY', TRUE),
            ('CN_FUND_002',    'CN Fund B',     'CNY', TRUE),
            ('US_STK_AAPL',   'Apple',         'USD', TRUE),
            ('INS_001',        'Insurance',     'CNY', TRUE)
        ON CONFLICT (canonical_id) DO NOTHING
    """)

    today = "2026-06-12"

    # CN Fund holdings
    conn.execute("""
        INSERT INTO holdings
            (asset_id, asset_name, source_system, snapshot_date,
             quantity, market_price_unit, market_value, currency, is_shadow)
        VALUES
            ('CN_FUND_001', 'CN Fund A', 'CN_Fund_Excel', ?, 1000, 1.5, 1500.0, 'CNY', FALSE),
            ('CN_FUND_002', 'CN Fund B', 'CN_Fund_Excel', ?, 2000, 2.0, 4000.0, 'CNY', FALSE)
    """, [today, today])

    # Schwab USD holding
    conn.execute("""
        INSERT INTO holdings
            (asset_id, asset_name, source_system, snapshot_date,
             quantity, market_price_unit, market_value, currency, is_shadow)
        VALUES
            ('US_STK_AAPL', 'Apple', 'Schwab_CSV', ?, 10, 200.0, 13524.2, 'USD', FALSE)
    """, [today])

    # RSU AMZN (USD)
    conn.execute("""
        INSERT INTO holdings
            (asset_id, asset_name, source_system, snapshot_date,
             quantity, market_price_unit, market_value, currency, is_shadow)
        VALUES
            ('RSU_AMZN', 'Amazon RSU', 'RSU_Excel', ?, 100, 250.0, 168500.0, 'USD', FALSE)
    """, [today])

    # Paper gold (CNY)
    conn.execute("""
        INSERT INTO holdings
            (asset_id, asset_name, source_system, snapshot_date,
             quantity, market_price_unit, market_value, currency, is_shadow)
        VALUES
            ('ALTS_Paper_Gold', 'Paper Gold', 'Gold_Excel', ?, 500.0, 500.0, 250000.0, 'CNY', FALSE)
    """, [today])

    # Insurance (CNY)
    conn.execute("""
        INSERT INTO holdings
            (asset_id, asset_name, source_system, snapshot_date,
             quantity, market_price_unit, market_value, currency, is_shadow)
        VALUES
            ('INS_001', 'Insurance', 'Insurance_Excel', ?, 1, 41713.0, 41713.0, 'CNY', FALSE)
    """, [today])

    # market_daily: AMZN, GOOGL prices
    conn.execute("""
        INSERT INTO market_daily (code, date, close, data_source)
        VALUES
            ('AMZN',  ?, 250.0, 'yfinance'),
            ('GOOGL', ?, 360.0, 'yfinance')
    """, [today, today])

    yield conn
    conn.close()


@pytest.fixture
def minimal_config(tmp_path) -> Dict[str, Any]:
    return {"finance_dir": str(tmp_path)}


@pytest.fixture
def ref_config_yaml(tmp_path) -> Path:
    """Write a minimal reference_sheet.yaml to tmp_path and return the path."""
    content = textwrap.dedent("""
        items:
          - fs_label: "投资资产_股票基金_A股基金"
            source:
              type: source_sum
              source_systems: ["CN_Fund_Excel"]
            emit:
              cny: true

          - fs_label: "投资资产_股票基金_美股基金_Schwab"
            source:
              type: source_sum
              source_systems: ["Schwab_CSV"]
            emit:
              cny: true
              usd: true

          - fs_label: "投资资产_股票基金_美股基金_IBKR"
            source:
              type: source_sum
              source_systems: ["Broker_IBKR"]
            emit:
              cny: true
              usd: true

          - fs_label: "投资资产_公司RSU_Amazon Stock"
            source:
              type: asset
              asset_id: "RSU_AMZN"
            emit:
              cny: true
              usd: true

          - fs_label: "投资资产_公司RSU_Google Stock"
            source:
              type: asset
              asset_id: "RSU_GOOGL"
            emit:
              cny: true
              usd: true

          - fs_label: "投资资产_黄金_纸黄金(元)"
            source:
              type: asset
              asset_id: "ALTS_Paper_Gold"
            emit:
              cny: true

          - fs_label: "投资资产_黄金_纸黄金(克)"
            source:
              type: asset
              asset_id: "ALTS_Paper_Gold"
            emit:
              qty: true

          - fs_label: "投资资产_长期保险_安泰人生"
            source:
              type: source_sum
              source_systems: ["Insurance_Excel"]
            emit:
              cny: true

          - fs_label: "USD Rate"
            source:
              type: price_row
              price_type: "fx_usdcny"
            emit:
              fx: true

          - fs_label: "Amazon Stock Price"
            source:
              type: price_row
              price_type: "market_price"
              ticker: "AMZN"
            emit:
              price: true

          - fs_label: "Google Stock Price"
            source:
              type: price_row
              price_type: "market_price"
              ticker: "GOOGL"
            emit:
              price: true
    """)
    p = tmp_path / "reference_sheet.yaml"
    p.write_text(content, encoding="utf-8")
    return p


# ── Safety guard tests ────────────────────────────────────────────────────────

class TestSafetyGuard:
    def test_correct_filename_passes(self, tmp_path):
        _assert_safe_output_path(tmp_path / OUTPUT_FILENAME)

    def test_wrong_filename_rejected(self, tmp_path):
        with pytest.raises(ValueError, match="filename must be"):
            _assert_safe_output_path(tmp_path / "Financial Summary_new.xlsx")

    def test_financial_summary_stem_rejected(self, tmp_path):
        """Even if the name happens to be the correct output name, reject protected stems."""
        # construct a path whose stem matches a forbidden name
        with pytest.raises(ValueError):
            _assert_safe_output_path(tmp_path / "funding_transactions.xlsx")

    def test_non_uis_reference_filename_rejected(self, tmp_path):
        with pytest.raises(ValueError, match="filename must be"):
            _assert_safe_output_path(tmp_path / "random_export.xlsx")


# ── Config loader tests ───────────────────────────────────────────────────────

class TestConfigLoader:
    def test_loads_items_from_yaml(self, ref_config_yaml):
        items = _load_reference_config(ref_config_yaml)
        assert len(items) == 11  # 美股基金 is now split into _Schwab and _IBKR rows
        labels = [i["fs_label"] for i in items]
        assert "投资资产_股票基金_A股基金" in labels
        assert "投资资产_股票基金_美股基金_Schwab" in labels
        assert "投资资产_股票基金_美股基金_IBKR" in labels
        assert "USD Rate" in labels

    def test_missing_config_raises_file_not_found(self, tmp_path):
        with pytest.raises(FileNotFoundError):
            _load_reference_config(tmp_path / "nonexistent.yaml")

    def test_config_order_preserved(self, ref_config_yaml):
        """Config item order must be stable (external links rely on row positions)."""
        items = _load_reference_config(ref_config_yaml)
        labels = [i["fs_label"] for i in items]
        # A股基金 must come before USD Rate
        assert labels.index("投资资产_股票基金_A股基金") < labels.index("USD Rate")

    def test_falls_back_to_example_template_when_real_file_missing(self, tmp_path):
        """Program OSR WS-4b: a real reference_sheet.yaml missing but a
        committed .example twin present must load from the example."""
        example = tmp_path / "reference_sheet.example.yaml"
        example.write_text(
            "items:\n"
            "  - fs_label: \"Example Row\"\n"
            "    source:\n"
            "      type: price_row\n"
            "      price_type: \"fx_usdcny\"\n"
            "    emit:\n"
            "      fx: true\n",
            encoding="utf-8",
        )
        real_path = tmp_path / "reference_sheet.yaml"  # deliberately not created

        items = _load_reference_config(real_path)

        assert len(items) == 1
        assert items[0]["fs_label"] == "Example Row"

    def test_real_file_wins_over_example_when_both_present(self, tmp_path):
        example = tmp_path / "reference_sheet.example.yaml"
        example.write_text(
            "items:\n  - fs_label: \"From Example\"\n    source: {type: price_row, price_type: fx_usdcny}\n"
            "    emit: {fx: true}\n",
            encoding="utf-8",
        )
        real = tmp_path / "reference_sheet.yaml"
        real.write_text(
            "items:\n  - fs_label: \"From Real\"\n    source: {type: price_row, price_type: fx_usdcny}\n"
            "    emit: {fx: true}\n",
            encoding="utf-8",
        )

        items = _load_reference_config(real)

        assert items[0]["fs_label"] == "From Real"

    def test_committed_example_template_loads(self):
        """config/reference_sheet.example.yaml as committed must parse and
        contain the persona insurance label (Program OSR WS-4b/WS-5b)."""
        items = _load_reference_config(Path("config/reference_sheet.example.yaml"))
        labels = [i["fs_label"] for i in items]
        assert "投资资产_长期保险_安泰人生" in labels


# ── Query helper tests ────────────────────────────────────────────────────────

class TestQueryHelpers:
    def test_source_sum_cny(self, mem_db):
        cny, usd, as_of = _query_source_sum(mem_db, ["CN_Fund_Excel"])
        assert cny == pytest.approx(5500.0)  # 1500 + 4000
        assert usd is None  # CN Fund is CNY, no USD column
        assert as_of == "2026-06-12"

    def test_source_sum_usd(self, mem_db):
        """Schwab USD holding: usd_total = qty * market_price_unit."""
        cny, usd, as_of = _query_source_sum(mem_db, ["Schwab_CSV"])
        assert cny == pytest.approx(13524.2)
        assert usd == pytest.approx(10 * 200.0)  # 2000.0

    def test_source_sum_multiple_sources(self, mem_db):
        cny, usd, as_of = _query_source_sum(mem_db, ["CN_Fund_Excel", "Schwab_CSV"])
        assert cny == pytest.approx(5500.0 + 13524.2)

    def test_asset_rsu_amzn(self, mem_db):
        cny, usd, qty, price, as_of = _query_asset(mem_db, "RSU_AMZN")
        assert cny == pytest.approx(168500.0)
        assert usd == pytest.approx(100 * 250.0)  # 25000.0
        assert qty == pytest.approx(100.0)
        assert price == pytest.approx(250.0)

    def test_asset_missing_returns_none(self, mem_db):
        """RSU_GOOGL not in fixture → all None (blank row, not an error)."""
        cny, usd, qty, price, as_of = _query_asset(mem_db, "RSU_GOOGL")
        assert cny is None
        assert usd is None
        assert qty is None

    def test_asset_gold_cny_qty(self, mem_db):
        cny, usd, qty, price, as_of = _query_asset(mem_db, "ALTS_Paper_Gold")
        assert cny == pytest.approx(250000.0)
        assert qty == pytest.approx(500.0)
        # Gold is CNY so usd should be None (currency != 'USD')
        assert usd is None

    def test_fx_rate_from_currency_service(self, mem_db):
        """_get_fx_rate reads from the live currency service (same source as Dashboard).
        The rate must be plausible (6 < rate < 8).
        as_of is always None (service does not return a date).
        """
        mock_service = MagicMock()
        mock_service.get_latest_rate.return_value = 7.2345
        with patch("src.sync.reference_export.get_currency_service", return_value=mock_service):
            rate, as_of = _get_fx_rate(mem_db)
        assert rate == pytest.approx(7.2345)
        assert as_of is None

    def test_fx_rate_fallback_on_none(self, mem_db):
        """If currency service returns None, fall back to 7.0."""
        mock_service = MagicMock()
        mock_service.get_latest_rate.return_value = None
        with patch("src.sync.reference_export.get_currency_service", return_value=mock_service):
            rate, as_of = _get_fx_rate(mem_db)
        assert rate == pytest.approx(7.0)
        assert as_of is None

    def test_fx_rate_fallback_on_exception(self, mem_db):
        """If currency service raises, fall back to 7.0."""
        with patch(
            "src.sync.reference_export.get_currency_service",
            side_effect=RuntimeError("service unavailable"),
        ):
            rate, as_of = _get_fx_rate(mem_db)
        assert rate == pytest.approx(7.0)
        assert as_of is None

    def test_market_price_amzn(self, mem_db):
        price, as_of = _get_market_price(mem_db, "AMZN")
        assert price == pytest.approx(250.0)
        assert as_of == "2026-06-12"

    def test_market_price_missing_ticker(self, mem_db):
        price, as_of = _get_market_price(mem_db, "MISSING")
        assert price is None
        assert as_of is None


# ── Row builder tests ─────────────────────────────────────────────────────────

class TestBuildRows:
    def test_rows_order_matches_config(self, mem_db, ref_config_yaml):
        items = _load_reference_config(ref_config_yaml)
        rows = _build_rows(mem_db, items)
        assert len(rows) == len(items)
        labels = [r[0] for r in rows]
        assert labels == [i["fs_label"] for i in items]

    def test_cn_fund_row_cny_sum(self, mem_db, ref_config_yaml):
        items = _load_reference_config(ref_config_yaml)
        rows = _build_rows(mem_db, items)
        row_map = {r[0]: r for r in rows}
        cn_row = row_map["投资资产_股票基金_A股基金"]
        # [label, cny, usd, qty, price, as_of]
        assert cn_row[1] == pytest.approx(5500.0)
        assert cn_row[2] is None  # usd not requested for A股基金

    def test_schwab_row_cny_and_usd(self, mem_db, ref_config_yaml):
        items = _load_reference_config(ref_config_yaml)
        rows = _build_rows(mem_db, items)
        row_map = {r[0]: r for r in rows}
        sw_row = row_map["投资资产_股票基金_美股基金_Schwab"]
        assert sw_row[1] == pytest.approx(13524.2)
        assert sw_row[2] == pytest.approx(2000.0)

    def test_ibkr_row_blank_when_no_data(self, mem_db, ref_config_yaml):
        """IBKR row is blank (None) until Workstream C goes live."""
        items = _load_reference_config(ref_config_yaml)
        rows = _build_rows(mem_db, items)
        row_map = {r[0]: r for r in rows}
        ibkr_row = row_map["投资资产_股票基金_美股基金_IBKR"]
        assert ibkr_row[1] is None  # cny — no IBKR data yet
        assert ibkr_row[2] is None  # usd

    def test_google_rsu_blank_when_missing(self, mem_db, ref_config_yaml):
        """RSU_GOOGL absent → row exists but all value cells are None (not an error)."""
        items = _load_reference_config(ref_config_yaml)
        rows = _build_rows(mem_db, items)
        row_map = {r[0]: r for r in rows}
        gg_row = row_map["投资资产_公司RSU_Google Stock"]
        assert gg_row[1] is None
        assert gg_row[2] is None

    def test_gold_qty_row(self, mem_db, ref_config_yaml):
        items = _load_reference_config(ref_config_yaml)
        rows = _build_rows(mem_db, items)
        row_map = {r[0]: r for r in rows}
        gold_qty_row = row_map["投资资产_黄金_纸黄金(克)"]
        assert gold_qty_row[3] == pytest.approx(500.0)  # qty column

    def test_usd_rate_row(self, mem_db, ref_config_yaml):
        items = _load_reference_config(ref_config_yaml)
        mock_service = MagicMock()
        mock_service.get_latest_rate.return_value = 7.15
        with patch("src.sync.reference_export.get_currency_service", return_value=mock_service):
            rows = _build_rows(mem_db, items)
        row_map = {r[0]: r for r in rows}
        fx_row = row_map["USD Rate"]
        # price column (index 4) holds the FX rate
        rate = fx_row[4]
        assert rate == pytest.approx(7.15)

    def test_amazon_price_row(self, mem_db, ref_config_yaml):
        items = _load_reference_config(ref_config_yaml)
        rows = _build_rows(mem_db, items)
        row_map = {r[0]: r for r in rows}
        amzn_row = row_map["Amazon Stock Price"]
        assert amzn_row[4] == pytest.approx(250.0)


# ── Full export tests ─────────────────────────────────────────────────────────

class TestExportReferenceSheet:
    def test_writes_uis_reference_data_xlsx(self, mem_db, minimal_config, ref_config_yaml):
        out_path = export_reference_sheet(
            mem_db, minimal_config, out_dir=Path(minimal_config["finance_dir"]),
            config_path=ref_config_yaml,
        )
        assert out_path.name == OUTPUT_FILENAME
        assert out_path.exists()

    def test_sheet_named_reference(self, mem_db, minimal_config, ref_config_yaml, tmp_path):
        import openpyxl
        out_path = export_reference_sheet(
            mem_db, minimal_config, out_dir=tmp_path, config_path=ref_config_yaml,
        )
        wb = openpyxl.load_workbook(str(out_path))
        assert "Reference" in wb.sheetnames

    def test_sheet_has_expected_labels_in_config_order(
        self, mem_db, minimal_config, ref_config_yaml, tmp_path
    ):
        import openpyxl
        out_path = export_reference_sheet(
            mem_db, minimal_config, out_dir=tmp_path, config_path=ref_config_yaml,
        )
        wb = openpyxl.load_workbook(str(out_path))
        ws = wb["Reference"]

        # Row 1 = "As of …" banner; Row 2 = header; Row 3+ = data
        data_labels = [ws.cell(row=i, column=1).value for i in range(3, ws.max_row + 1)]
        items = _load_reference_config(ref_config_yaml)
        expected_labels = [item["fs_label"] for item in items]
        assert data_labels == expected_labels

    def test_cn_fund_cny_value_in_sheet(
        self, mem_db, minimal_config, ref_config_yaml, tmp_path
    ):
        import openpyxl
        out_path = export_reference_sheet(
            mem_db, minimal_config, out_dir=tmp_path, config_path=ref_config_yaml,
        )
        wb = openpyxl.load_workbook(str(out_path))
        ws = wb["Reference"]

        # Find the A股基金 row
        for row in ws.iter_rows(min_row=3, values_only=True):
            if row[0] == "投资资产_股票基金_A股基金":
                cny_val = row[1]
                assert cny_val == pytest.approx(5500.0)
                return
        pytest.fail("A股基金 label not found in sheet")

    def test_safety_guard_prevents_wrong_filename(self, mem_db, tmp_path, ref_config_yaml):
        """Passing a directory whose joining would yield a wrong filename triggers guard."""
        # Patch OUTPUT_FILENAME temporarily to see the guard catch wrong name
        import src.sync.reference_export as ref_mod
        original = ref_mod.OUTPUT_FILENAME
        ref_mod.OUTPUT_FILENAME = "UIS_Reference_Data.xlsx"  # correct; now test wrong path
        try:
            with pytest.raises(ValueError, match="Safety guard"):
                _assert_safe_output_path(tmp_path / "financial summary_new.xlsx")
        finally:
            ref_mod.OUTPUT_FILENAME = original

    def test_missing_finance_dir_raises(self, mem_db, ref_config_yaml):
        """No out_dir and no finance_dir in config → FileNotFoundError."""
        with pytest.raises(FileNotFoundError):
            export_reference_sheet(
                mem_db, {}, out_dir=None, config_path=ref_config_yaml,
            )

    def test_creates_output_dir_if_missing(self, mem_db, tmp_path, ref_config_yaml):
        """Output dir that doesn't exist should be created."""
        nested = tmp_path / "new" / "subdir"
        assert not nested.exists()
        out_path = export_reference_sheet(
            mem_db, {}, out_dir=nested, config_path=ref_config_yaml,
        )
        assert out_path.exists()

    def test_dry_run_does_not_call_export(self, mem_db):
        """orchestrator dry_run=True should skip the reference export."""
        from unittest.mock import patch as _patch
        from src.sync.orchestrator import run_full_sync_v3

        called = []
        def fake_export(connector, config, **kw):
            called.append(True)
            return Path("/tmp/UIS_Reference_Data.xlsx")

        with _patch("src.sync.orchestrator.sync_current_allocations", return_value={"synced": 0}), \
             _patch("src.sync.orchestrator.validate_cost_basis", return_value=[]), \
             _patch("src.sync.orchestrator.validate_allocations", return_value=[]):
            # Import and patch reference_export inside the orchestrator namespace
            import src.sync.reference_export as ref_mod
            with _patch.object(ref_mod, "export_reference_sheet", side_effect=fake_export):
                run_full_sync_v3(mem_db, {}, dry_run=True)

        assert called == [], "export_reference_sheet must NOT be called during dry_run"


# ── Config extensibility test ─────────────────────────────────────────────────

class TestConfigExtensibility:
    """Prove that adding a new item to the YAML produces a new row (GSU pattern)."""

    def test_adding_gsu_item_yields_new_row(self, mem_db, tmp_path):
        # Seed a RSU_GOOGL holding so the new item has data
        conn = mem_db
        conn.execute("""
            INSERT INTO holdings
                (asset_id, asset_name, source_system, snapshot_date,
                 quantity, market_price_unit, market_value, currency, is_shadow)
            VALUES
                ('RSU_GOOGL', 'Google RSU', 'RSU_Excel', '2026-06-12',
                 50, 360.0, 121500.0, 'USD', FALSE)
        """)

        # Write a config with just CN Fund + the new GSU row
        config_content = textwrap.dedent("""
            items:
              - fs_label: "投资资产_股票基金_A股基金"
                source:
                  type: source_sum
                  source_systems: ["CN_Fund_Excel"]
                emit:
                  cny: true
              - fs_label: "投资资产_公司RSU_Google Stock"
                source:
                  type: asset
                  asset_id: "RSU_GOOGL"
                emit:
                  cny: true
                  usd: true
        """)
        cfg_path = tmp_path / "ref_gsu.yaml"
        cfg_path.write_text(config_content, encoding="utf-8")

        out_path = export_reference_sheet(conn, {}, out_dir=tmp_path, config_path=cfg_path)

        import openpyxl
        wb = openpyxl.load_workbook(str(out_path))
        ws = wb["Reference"]
        labels = [ws.cell(row=i, column=1).value for i in range(3, ws.max_row + 1)]
        assert "投资资产_公司RSU_Google Stock" in labels

        # Verify the GSU CNY value is populated
        for row in ws.iter_rows(min_row=3, values_only=True):
            if row[0] == "投资资产_公司RSU_Google Stock":
                assert row[1] == pytest.approx(121500.0)
                assert row[2] == pytest.approx(50 * 360.0)
                return
        pytest.fail("GSU row value not found in sheet")


# ── WS-2: co-authority totals (§ deliverable 1) ────────────────────────────────
# Schwab_CSV + Broker_IBKR are real co-authority sources in config/source_authority.yaml
# (US_STK_*, US_ETF_*, CASH_USD patterns) — these tests rely on that real config, exactly
# like `test_ibkr_row_blank_when_no_data` above already implicitly does.

class TestCoauthorityShare:
    @pytest.fixture
    def coauth_db(self, mem_db):
        """mem_db + a co-authority SGOV position merged into a Consolidated row.

        IBKR contributed 100 units / $10,000 (mv 100000 CNY equiv at rate 10 for round
        numbers), Schwab contributed 200 units / $20,000 (mv 200000). Both broker rows are
        already shadowed (is_shadow=TRUE) — the state C3.4 leaves them in — and the merged
        Consolidated row is active (is_shadow=FALSE). IBKR also independently holds a
        non-co-authority... actually holds a SECOND, uniquely-IBKR asset that was never
        consolidated, to prove the share is ADDED to (not a replacement for) the broker's
        own direct total.
        """
        conn = mem_db
        today = "2026-06-12"
        conn.execute(
            """
            INSERT INTO holdings
                (asset_id, asset_name, source_system, snapshot_date,
                 quantity, market_price_unit, market_value, currency, is_shadow)
            VALUES
                ('US_ETF_SGOV', 'SGOV', 'Broker_IBKR', ?, 100, 100.0, 100000.0, 'USD', TRUE),
                ('US_ETF_SGOV', 'SGOV', 'Schwab_CSV',  ?, 200, 100.0, 200000.0, 'USD', TRUE),
                ('US_ETF_SGOV', 'SGOV', 'Consolidated', ?, 300, 100.0, 300000.0, 'USD', FALSE),
                ('US_STK_XYZ', 'XYZ Corp', 'Broker_IBKR', ?, 10, 500.0, 5000.0, 'USD', FALSE)
            """,
            [today, today, today, today],
        )
        return conn

    def test_broker_consolidated_share_recovers_ibkr(self, coauth_db):
        cny, usd, as_of = _query_broker_consolidated_share(coauth_db, "Broker_IBKR")
        assert cny == pytest.approx(100000.0)
        assert usd == pytest.approx(100 * 100.0)
        assert as_of == "2026-06-12"

    def test_broker_consolidated_share_recovers_schwab(self, coauth_db):
        cny, usd, as_of = _query_broker_consolidated_share(coauth_db, "Schwab_CSV")
        assert cny == pytest.approx(200000.0)
        assert usd == pytest.approx(200 * 100.0)

    def test_broker_consolidated_share_none_when_no_consolidated_row(self, mem_db):
        """No Consolidated rows in the plain mem_db fixture → share is None, not 0."""
        cny, usd, as_of = _query_broker_consolidated_share(mem_db, "Broker_IBKR")
        assert cny is None
        assert usd is None

    def test_ibkr_source_sum_includes_consolidated_share_plus_own(self, coauth_db):
        """IBKR's total = its own uniquely-held asset (5000) + its SGOV share (100000)."""
        cny, usd, as_of = _query_source_sum(coauth_db, ["Broker_IBKR"])
        assert cny == pytest.approx(5000.0 + 100000.0)
        assert usd == pytest.approx(10 * 500.0 + 100 * 100.0)

    def test_schwab_source_sum_includes_consolidated_share(self, coauth_db):
        """mem_db already seeds a Schwab_CSV AAPL holding (13524.2) — Schwab's total here
        is that own row PLUS its SGOV share (200000), proving additive, not replaced."""
        cny, usd, as_of = _query_source_sum(coauth_db, ["Schwab_CSV"])
        assert cny == pytest.approx(13524.2 + 200000.0)
        assert usd == pytest.approx(10 * 200.0 + 200 * 100.0)

    def test_non_coauthority_source_unaffected(self, coauth_db):
        """CN_Fund_Excel is not a co-authority source — no share-recovery attempted,
        totals match the plain is_shadow=FALSE sum (regression guard on non-coauthority
        rows)."""
        cny, usd, as_of = _query_source_sum(coauth_db, ["CN_Fund_Excel"])
        assert cny == pytest.approx(5500.0)


# ── WS-2: 月度快照 monthly snapshot (§ deliverable 2) ──────────────────────────

class TestMonthlySnapshotCompute:
    @pytest.fixture
    def multi_month_db(self):
        conn = DatabaseConnector(":memory:")
        initialize_schema(conn)
        conn.execute(
            """
            INSERT INTO asset_registry (canonical_id, display_name, base_currency, is_active)
            VALUES ('CN_FUND_001', 'CN Fund A', 'CNY', TRUE)
            ON CONFLICT (canonical_id) DO NOTHING
            """
        )
        # Three snapshots for the same asset across two months + one stale earlier snapshot
        # in the same month that must be ignored in favor of the per-asset month-end MAX.
        conn.execute(
            """
            INSERT INTO holdings
                (asset_id, asset_name, source_system, snapshot_date,
                 quantity, market_price_unit, market_value, currency, is_shadow)
            VALUES
                ('CN_FUND_001', 'CN Fund A', 'CN_Fund_Excel', '2026-01-05', 900, 1.0, 900.0, 'CNY', FALSE),
                ('CN_FUND_001', 'CN Fund A', 'CN_Fund_Excel', '2026-01-31', 1000, 1.1, 1100.0, 'CNY', FALSE),
                ('CN_FUND_001', 'CN Fund A', 'CN_Fund_Excel', '2026-02-15', 1200, 1.2, 1440.0, 'CNY', FALSE),
                ('CN_FUND_001', 'CN Fund A', 'CN_Fund_Excel', '2026-02-10', 1100, 1.15, 1265.0, 'CNY', FALSE)
            """
        )
        yield conn
        conn.close()

    def test_picks_per_asset_month_end_max(self, multi_month_db):
        rows = _compute_monthly_snapshot_rows(multi_month_db, "2026-01", "2026-02")
        by_month = {r[0]: r for r in rows}
        assert by_month["2026-01"][4] == pytest.approx(1000.0)  # qty from 01-31, not 01-05
        assert by_month["2026-01"][5] == pytest.approx(1100.0)
        assert by_month["2026-02"][4] == pytest.approx(1200.0)  # qty from 02-15, not 02-10
        assert by_month["2026-02"][5] == pytest.approx(1440.0)

    def test_row_shape(self, multi_month_db):
        rows = _compute_monthly_snapshot_rows(multi_month_db, "2026-01", "2026-02")
        assert len(rows) == 2
        month, asset_id, asset_name, source, qty, mv = rows[0]
        assert asset_id == "CN_FUND_001"
        assert source == "CN_Fund_Excel"

    def test_excludes_out_of_range_months(self, multi_month_db):
        rows = _compute_monthly_snapshot_rows(multi_month_db, "2026-02", "2026-02")
        assert len(rows) == 1
        assert rows[0][0] == "2026-02"

    def test_excludes_shadowed_rows(self):
        conn = DatabaseConnector(":memory:")
        initialize_schema(conn)
        conn.execute(
            """
            INSERT INTO holdings
                (asset_id, asset_name, source_system, snapshot_date,
                 quantity, market_price_unit, market_value, currency, is_shadow)
            VALUES ('CN_FUND_001', 'CN Fund A', 'CN_Fund_Excel', '2026-03-31',
                    500, 1.0, 500.0, 'CNY', TRUE)
            """
        )
        rows = _compute_monthly_snapshot_rows(conn, "2026-03", "2026-03")
        assert rows == []
        conn.close()


class TestMonthlySnapshotMerge:
    def test_append_only_freezes_closed_month(self):
        """A pre-existing row for a CLOSED month survives a re-export even when the DB
        would now compute a different value for that same (month, asset_id)."""
        existing = {
            ("2026-01", "CN_FUND_001"): ["2026-01", "CN_FUND_001", "CN Fund A", "CN_Fund_Excel", 1000.0, 1100.0],
        }
        computed = [
            ("2026-01", "CN_FUND_001", "CN Fund A", "CN_Fund_Excel", 9999.0, 8888.0),  # would-be new value
            ("2026-02", "CN_FUND_001", "CN Fund A", "CN_Fund_Excel", 1200.0, 1440.0),
        ]
        merged = _merge_monthly_snapshot(existing, computed, current_month="2026-02")
        jan_row = next(r for r in merged if r[0] == "2026-01")
        assert jan_row[4] == 1000.0  # frozen — NOT overwritten by computed's 9999.0
        assert jan_row[5] == 1100.0

    def test_current_month_always_refreshed(self):
        """The open (current) month is replaced by computed values even if a stale
        pre-existing row for it is present."""
        existing = {
            ("2026-02", "CN_FUND_001"): ["2026-02", "CN_FUND_001", "CN Fund A", "CN_Fund_Excel", 1.0, 1.0],
        }
        computed = [
            ("2026-02", "CN_FUND_001", "CN Fund A", "CN_Fund_Excel", 1200.0, 1440.0),
        ]
        merged = _merge_monthly_snapshot(existing, computed, current_month="2026-02")
        assert len(merged) == 1
        assert merged[0][4] == 1200.0
        assert merged[0][5] == 1440.0

    def test_first_time_backfill_fills_all_closed_months(self):
        """No pre-existing sheet (first-ever export) → all computed rows, including
        closed months, are written."""
        merged = _merge_monthly_snapshot(
            existing={},
            computed=[
                ("2026-01", "CN_FUND_001", "CN Fund A", "CN_Fund_Excel", 1000.0, 1100.0),
                ("2026-02", "CN_FUND_001", "CN Fund A", "CN_Fund_Excel", 1200.0, 1440.0),
            ],
            current_month="2026-02",
        )
        assert len(merged) == 2

    def test_newly_closed_month_gap_filled_once(self):
        """A month that has just closed since the last export (present in `computed` but
        not yet in `existing`) is added — then frozen on the NEXT merge."""
        existing = {
            ("2026-01", "CN_FUND_001"): ["2026-01", "CN_FUND_001", "CN Fund A", "CN_Fund_Excel", 1000.0, 1100.0],
        }
        computed = [
            ("2026-01", "CN_FUND_001", "CN Fund A", "CN_Fund_Excel", 9999.0, 8888.0),
            ("2026-02", "CN_FUND_001", "CN Fund A", "CN_Fund_Excel", 1200.0, 1440.0),
        ]
        # First merge: now-current month is 2026-03, so both Jan and Feb are closed.
        merged = _merge_monthly_snapshot(existing, computed, current_month="2026-03")
        by_month = {r[0]: r for r in merged}
        assert by_month["2026-01"][4] == 1000.0  # still frozen from existing
        assert by_month["2026-02"][4] == 1200.0  # gap-filled from computed (first time seen)


class TestMonthlySnapshotSheetRoundTrip:
    def test_sheet_written_and_reloadable(self, mem_db, minimal_config, ref_config_yaml, tmp_path):
        import openpyxl

        out_path = export_reference_sheet(
            mem_db, minimal_config, out_dir=Path(minimal_config["finance_dir"]),
            config_path=ref_config_yaml,
        )
        wb = openpyxl.load_workbook(str(out_path))
        assert MONTHLY_SHEET_NAME in wb.sheetnames
        ws = wb[MONTHLY_SHEET_NAME]
        header = [ws.cell(row=1, column=c).value for c in range(1, 7)]
        assert header == ["Month", "Asset_ID", "Asset_Name", "Source", "Qty", "Market_Value_CNY"]
        # mem_db fixture data is all dated 2026-06-12 → one row per asset for month 2026-06.
        data_rows = list(ws.iter_rows(min_row=2, values_only=True))
        assert len(data_rows) >= 1
        assert all(r[0] == "2026-06" for r in data_rows)

    def test_reexport_freezes_prior_month(self, mem_db, minimal_config, ref_config_yaml, tmp_path):
        """Simulate: an export already ran and wrote a row for a now-closed month with a
        DIFFERENT value than the DB currently has. A re-export must not rewrite it."""
        out_path = Path(minimal_config["finance_dir"]) / OUTPUT_FILENAME

        # Seed a pre-existing workbook with a frozen January row that intentionally
        # diverges from what `mem_db` would compute for CN_FUND_001 (which is dated
        # 2026-06-12, so it wouldn't even produce a January row — this proves frozen
        # history survives independently of what the live DB currently contains).
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reference"
        ws.append(["As of 2026-01-31 00:00"])
        ws.append(["Label", "Value_CNY", "Value_USD", "Qty", "Price", "AsOf"])
        ws_monthly = wb.create_sheet(title=MONTHLY_SHEET_NAME)
        ws_monthly.append(["Month", "Asset_ID", "Asset_Name", "Source", "Qty", "Market_Value_CNY"])
        ws_monthly.append(["2026-01", "CN_FUND_001", "CN Fund A", "CN_Fund_Excel", 42.0, 4242.0])
        wb.save(str(out_path))

        export_reference_sheet(
            mem_db, minimal_config, out_dir=Path(minimal_config["finance_dir"]),
            config_path=ref_config_yaml,
        )

        wb2 = openpyxl.load_workbook(str(out_path))
        ws2 = wb2[MONTHLY_SHEET_NAME]
        rows = {(r[0], r[1]): r for r in ws2.iter_rows(min_row=2, values_only=True)}
        frozen = rows[("2026-01", "CN_FUND_001")]
        assert frozen[4] == 42.0
        assert frozen[5] == 4242.0
