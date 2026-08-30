"""Tests for CN Fund raw transaction/holdings processor.

TDD: Tests written FIRST before implementation.
Run: pytest tests/sources/test_cn_fund_raw_processor.py -v

The processor:
1. Reads raw_transactions_paste, finds rows not in 基金交易记录, appends them
2. Reads raw_holdings_paste, aggregates by fund, finds new snapshots, appends them
"""
import io
import pytest

pytestmark = pytest.mark.pipeline

from datetime import datetime, date
import openpyxl


# ---------------------------------------------------------------------------
# Helpers: build in-memory workbooks
# ---------------------------------------------------------------------------

def _make_workbook_bytes(sheets: dict) -> bytes:
    """Create an in-memory xlsx workbook with given sheet data.

    sheets: {sheet_name: [header_tuple, row_tuple, ...]}
    Returns bytes of the xlsx file.
    """
    wb = openpyxl.Workbook()
    first = True
    for sheet_name, rows in sheets.items():
        if first:
            ws = wb.active
            ws.title = sheet_name
            first = False
        else:
            ws = wb.create_sheet(sheet_name)
        for row in rows:
            ws.append(list(row))
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _workbook_sheet_rows(wb_bytes: bytes, sheet_name: str):
    """Read all rows (as tuples) from a sheet in a workbook byte string."""
    wb = openpyxl.load_workbook(io.BytesIO(wb_bytes), data_only=True)
    ws = wb[sheet_name]
    return [tuple(c.value for c in row) for row in ws.iter_rows()]


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

RAW_TX_HEADER = ('确认日期', '基金代码', '基金简称', '业务类型', '确认状态', '确认份额', '确认金额', '手续费', '确认净值', '关联银行卡')
PROCESSED_TX_HEADER = ('交易日期', '基金代码', '基金名称', '操作类型', '交易金额', '交易份额', '交易时基金单位净值', '手续费', '交易原因')

RAW_HOLD_HEADER = ('基金代码', '基金简称', '基金类型', '净值日期', '单位净值', '持有份额', '参考市值', '关联银行卡')
PROCESSED_HOLD_HEADER = ('Asset_ID', 'Asset_Name', 'Asset_Type_Raw', 'Snapshot_Date', 'Market_Price_Unit', 'Quantity', 'Market_Value_Raw')


@pytest.fixture
def workbook_with_new_dividend(tmp_path):
    """Workbook where raw_transactions_paste has a 现金分红 not in 基金交易记录."""
    wb_bytes = _make_workbook_bytes({
        'raw_transactions_paste': [
            RAW_TX_HEADER,
            (datetime(2026, 2, 10), '900012', '示例中证红利指数增强A', '现金分红', '成功', None, 3323.2, None, None, '工行|1234'),
        ],
        'raw_holdings_paste': [
            RAW_HOLD_HEADER,
        ],
        '基金持仓汇总': [
            PROCESSED_HOLD_HEADER,
        ],
        '基金交易记录': [
            PROCESSED_TX_HEADER,
            # Existing row for a different fund — should not deduplicate our new row
            (datetime(2026, 1, 22), '4179', '示例稳健货币B', '申购', 38869.17, 38869.17, 1.0, 0.0, '手动买入'),
        ],
    })
    p = tmp_path / 'funding_transactions.xlsx'
    p.write_bytes(wb_bytes)
    return p


@pytest.fixture
def workbook_already_processed(tmp_path):
    """Workbook where the dividend is already in 基金交易记录 — should be a no-op."""
    wb_bytes = _make_workbook_bytes({
        'raw_transactions_paste': [
            RAW_TX_HEADER,
            (datetime(2026, 2, 10), '900012', '示例中证红利指数增强A', '现金分红', '成功', None, 3323.2, None, None, '工行|1234'),
        ],
        'raw_holdings_paste': [RAW_HOLD_HEADER],
        '基金持仓汇总': [PROCESSED_HOLD_HEADER],
        '基金交易记录': [
            PROCESSED_TX_HEADER,
            # Same transaction already present
            (datetime(2026, 2, 10), '900012', '示例中证红利指数增强A', '现金分红', 3323.2, 0.0, None, None, 'Cash Dividend'),
        ],
    })
    p = tmp_path / 'funding_transactions.xlsx'
    p.write_bytes(wb_bytes)
    return p


@pytest.fixture
def workbook_failed_status(tmp_path):
    """Raw row with 确认状态='失败' — should be ignored."""
    wb_bytes = _make_workbook_bytes({
        'raw_transactions_paste': [
            RAW_TX_HEADER,
            (datetime(2026, 2, 10), '900012', '示例中证红利指数增强A', '现金分红', '失败', None, 3323.2, None, None, None),
        ],
        'raw_holdings_paste': [RAW_HOLD_HEADER],
        '基金持仓汇总': [PROCESSED_HOLD_HEADER],
        '基金交易记录': [PROCESSED_TX_HEADER],
    })
    p = tmp_path / 'funding_transactions.xlsx'
    p.write_bytes(wb_bytes)
    return p


@pytest.fixture
def workbook_type_mapping(tmp_path):
    """Various 业务类型 values to verify type mapping logic."""
    wb_bytes = _make_workbook_bytes({
        'raw_transactions_paste': [
            RAW_TX_HEADER,
            (datetime(2026, 2, 1), '900001', 'Fund A', '买入',    '成功', 1000.0, 1000.0, 1.5, 1.0, None),
            (datetime(2026, 2, 2), '900001', 'Fund A', '卖出',    '成功', 500.0,  500.0,  0.0, 1.0, None),
            (datetime(2026, 2, 3), '900001', 'Fund A', '定投',    '成功', 300.0,  300.0,  0.0, 1.0, None),
            (datetime(2026, 2, 4), '900001', 'Fund A', '快速取现', '成功', 200.0,  200.0,  0.0, 1.0, None),
            (datetime(2026, 2, 5), '900001', 'Fund A', '充值',    '成功', 100.0,  100.0,  0.0, 1.0, None),
            (datetime(2026, 2, 6), '900001', 'Fund A', '红利再投资', '成功', 50.0, 50.0,  0.0, 1.0, None),
        ],
        'raw_holdings_paste': [RAW_HOLD_HEADER],
        '基金持仓汇总': [PROCESSED_HOLD_HEADER],
        '基金交易记录': [PROCESSED_TX_HEADER],
    })
    p = tmp_path / 'funding_transactions.xlsx'
    p.write_bytes(wb_bytes)
    return p


@pytest.fixture
def workbook_with_new_holdings(tmp_path):
    """Workbook where raw_holdings_paste has a Feb snapshot not in 基金持仓汇总."""
    wb_bytes = _make_workbook_bytes({
        'raw_transactions_paste': [RAW_TX_HEADER],
        'raw_holdings_paste': [
            RAW_HOLD_HEADER,
            # Two rows for same fund (different bank accounts) → should aggregate
            (datetime(2026, 2, 27), '900001', '示例沪深300A', '指数型', datetime(2026, 2, 27), 2.756, 29963.97, 82580.7, '工行|1234'),
            (datetime(2026, 2, 27), '900001', '示例沪深300A', '指数型', datetime(2026, 2, 27), 2.756, 63893.21, 176089.69, '北京银行|9581'),
        ],
        '基金持仓汇总': [
            PROCESSED_HOLD_HEADER,
            # Only has Jan snapshot
            ('900001', '示例沪深300A', '指数型', datetime(2026, 1, 29), 2.65, 88000.0, 233200.0),
        ],
        '基金交易记录': [PROCESSED_TX_HEADER],
    })
    p = tmp_path / 'funding_transactions.xlsx'
    p.write_bytes(wb_bytes)
    return p


@pytest.fixture
def workbook_holdings_already_synced(tmp_path):
    """Holdings snapshot already in 基金持仓汇总 — should be a no-op."""
    wb_bytes = _make_workbook_bytes({
        'raw_transactions_paste': [RAW_TX_HEADER],
        'raw_holdings_paste': [
            RAW_HOLD_HEADER,
            (datetime(2026, 2, 27), '900001', '示例沪深300A', '指数型', datetime(2026, 2, 27), 2.756, 93857.18, 258727.39, '工行|1234'),
        ],
        '基金持仓汇总': [
            PROCESSED_HOLD_HEADER,
            # Feb 27 already present
            ('900001', '示例沪深300A', '指数型', datetime(2026, 2, 27), 2.756, 93857.18, 258727.39),
        ],
        '基金交易记录': [PROCESSED_TX_HEADER],
    })
    p = tmp_path / 'funding_transactions.xlsx'
    p.write_bytes(wb_bytes)
    return p


# ---------------------------------------------------------------------------
# Tests: process_raw_transactions
# ---------------------------------------------------------------------------

class TestProcessRawTransactions:

    def test_new_dividend_appended_to_processed_sheet(self, workbook_with_new_dividend):
        """A new 现金分红 in raw tab that isn't in 基金交易记录 gets appended."""
        from src.sources.cn_fund_raw_processor import process_raw_transactions

        result = process_raw_transactions(workbook_with_new_dividend)

        assert result.new_count == 1
        # Verify it was written to the file
        rows = _workbook_sheet_rows(workbook_with_new_dividend.read_bytes(), '基金交易记录')
        # Header + 1 existing + 1 new = 3 rows
        assert len(rows) == 3
        new_row = rows[2]
        assert new_row[1] == '900012'          # 基金代码
        assert new_row[3] == '现金分红'          # 操作类型 preserved
        assert new_row[4] == 3323.2            # 交易金额
        assert new_row[8] == 'Cash Dividend'   # 交易原因

    def test_already_processed_row_not_duplicated(self, workbook_already_processed):
        """Row already in 基金交易记录 is not appended again."""
        from src.sources.cn_fund_raw_processor import process_raw_transactions

        result = process_raw_transactions(workbook_already_processed)

        assert result.new_count == 0
        rows = _workbook_sheet_rows(workbook_already_processed.read_bytes(), '基金交易记录')
        assert len(rows) == 2  # header + 1 existing only

    def test_failed_status_row_skipped(self, workbook_failed_status):
        """Row with 确认状态='失败' is not appended."""
        from src.sources.cn_fund_raw_processor import process_raw_transactions

        result = process_raw_transactions(workbook_failed_status)

        assert result.new_count == 0
        rows = _workbook_sheet_rows(workbook_failed_status.read_bytes(), '基金交易记录')
        assert len(rows) == 1  # header only

    def test_type_mapping_buy_types(self, workbook_type_mapping):
        """买入 and 定投 map to 申购; 卖出 maps to 赎回."""
        from src.sources.cn_fund_raw_processor import process_raw_transactions

        result = process_raw_transactions(workbook_type_mapping)

        assert result.new_count == 6
        rows = _workbook_sheet_rows(workbook_type_mapping.read_bytes(), '基金交易记录')
        # Row 2 (index 1): 买入 → 申购
        assert rows[1][3] == '申购'
        # Row 3 (index 2): 卖出 → 赎回
        assert rows[2][3] == '赎回'
        # Row 4 (index 3): 定投 → 申购
        assert rows[3][3] == '申购'
        # Row 5 (index 4): 快速取现 → 快速取现 (passthrough)
        assert rows[4][3] == '快速取现'
        # Row 6 (index 5): 充值 → 活期宝即充即用
        assert rows[5][3] == '活期宝即充即用'
        # Row 7 (index 6): 红利再投资 → 红利再投资 (passthrough)
        assert rows[6][3] == '红利再投资'

    def test_date_and_fund_fields_mapped_correctly(self, workbook_with_new_dividend):
        """Date and fund fields land in the correct 基金交易记录 columns."""
        from src.sources.cn_fund_raw_processor import process_raw_transactions

        process_raw_transactions(workbook_with_new_dividend)

        rows = _workbook_sheet_rows(workbook_with_new_dividend.read_bytes(), '基金交易记录')
        new_row = rows[2]
        # 交易日期, 基金代码, 基金名称
        assert new_row[0].date() == date(2026, 2, 10) if hasattr(new_row[0], 'date') else new_row[0] == datetime(2026, 2, 10)
        assert new_row[1] == '900012'
        assert new_row[2] == '示例中证红利指数增强A'


# ---------------------------------------------------------------------------
# Tests: process_raw_holdings
# ---------------------------------------------------------------------------

class TestProcessRawHoldings:

    def test_new_snapshot_appended_aggregated(self, workbook_with_new_holdings):
        """Feb 27 snapshot aggregates two bank-account rows into one and appends it."""
        from src.sources.cn_fund_raw_processor import process_raw_holdings

        result = process_raw_holdings(workbook_with_new_holdings)

        assert result.new_count == 1
        rows = _workbook_sheet_rows(workbook_with_new_holdings.read_bytes(), '基金持仓汇总')
        # Header + 1 existing (Jan) + 1 new (Feb) = 3 rows
        assert len(rows) == 3
        new_row = rows[2]
        assert new_row[0] == '900001'                    # Asset_ID
        assert abs(new_row[5] - (29963.97 + 63893.21)) < 0.01  # Quantity aggregated
        assert abs(new_row[6] - (82580.7 + 176089.69)) < 0.01  # Market_Value aggregated

    def test_existing_snapshot_not_duplicated(self, workbook_holdings_already_synced):
        """Holdings snapshot already present is not added again."""
        from src.sources.cn_fund_raw_processor import process_raw_holdings

        result = process_raw_holdings(workbook_holdings_already_synced)

        assert result.new_count == 0
        rows = _workbook_sheet_rows(workbook_holdings_already_synced.read_bytes(), '基金持仓汇总')
        assert len(rows) == 2  # header + 1 existing only

    def test_snapshot_date_is_nav_date(self, workbook_with_new_holdings):
        """Snapshot_Date in 基金持仓汇总 comes from 净值日期 in raw_holdings_paste."""
        from src.sources.cn_fund_raw_processor import process_raw_holdings

        process_raw_holdings(workbook_with_new_holdings)

        rows = _workbook_sheet_rows(workbook_with_new_holdings.read_bytes(), '基金持仓汇总')
        new_row = rows[2]
        snap_date = new_row[3]
        assert snap_date.year == 2026
        assert snap_date.month == 2
        assert snap_date.day == 27


# ---------------------------------------------------------------------------
# Tests: process_all (combined entry point)
# ---------------------------------------------------------------------------

class TestProcessAll:

    def test_returns_combined_stats(self, workbook_with_new_dividend, workbook_with_new_holdings, tmp_path):
        """process_all returns new_transactions and new_holdings counts."""
        from src.sources.cn_fund_raw_processor import process_all

        # Build a single workbook with both new tx and new holdings
        wb_bytes = _make_workbook_bytes({
            'raw_transactions_paste': [
                RAW_TX_HEADER,
                (datetime(2026, 2, 10), '900012', '示例中证红利指数增强A', '现金分红', '成功', None, 3323.2, None, None, None),
            ],
            'raw_holdings_paste': [
                RAW_HOLD_HEADER,
                (datetime(2026, 2, 27), '900001', '示例沪深300A', '指数型', datetime(2026, 2, 27), 2.756, 93857.18, 258727.39, '工行|1234'),
            ],
            '基金持仓汇总': [PROCESSED_HOLD_HEADER],
            '基金交易记录': [PROCESSED_TX_HEADER],
        })
        p = tmp_path / 'combined.xlsx'
        p.write_bytes(wb_bytes)

        result = process_all(p)

        assert result.new_transactions == 1
        assert result.new_holdings == 1

    def test_no_new_data_is_noop(self, workbook_already_processed, workbook_holdings_already_synced, tmp_path):
        """process_all with nothing new returns zero counts and doesn't touch the file."""
        from src.sources.cn_fund_raw_processor import process_all

        wb_bytes = _make_workbook_bytes({
            'raw_transactions_paste': [
                RAW_TX_HEADER,
                (datetime(2026, 2, 10), '900012', '示例中证红利指数增强A', '现金分红', '成功', None, 3323.2, None, None, None),
            ],
            'raw_holdings_paste': [
                RAW_HOLD_HEADER,
                (datetime(2026, 2, 27), '900001', '示例沪深300A', '指数型', datetime(2026, 2, 27), 2.756, 93857.18, 258727.39, None),
            ],
            '基金持仓汇总': [
                PROCESSED_HOLD_HEADER,
                ('900001', '示例沪深300A', '指数型', datetime(2026, 2, 27), 2.756, 93857.18, 258727.39),
            ],
            '基金交易记录': [
                PROCESSED_TX_HEADER,
                (datetime(2026, 2, 10), '900012', '示例中证红利指数增强A', '现金分红', 3323.2, 0.0, None, None, 'Cash Dividend'),
            ],
        })
        p = tmp_path / 'noop.xlsx'
        p.write_bytes(wb_bytes)

        result = process_all(p)

        assert result.new_transactions == 0
        assert result.new_holdings == 0


# ---------------------------------------------------------------------------
# Tests: cn_fund_raw_process hook — GCS write-back (reader_hooks.py)
# ---------------------------------------------------------------------------

class TestCNFundRawProcessHookGCSWriteback:
    """Tests for the GCS write-back logic added to cn_fund_raw_process in V7.1.x.

    All tests mock process_all and upload_source_to_gcs — no real workbooks or
    network calls are made.  DatabaseConnector is never called.
    """

    @pytest.fixture()
    def fake_workbook(self, tmp_path):
        """A minimal placeholder file under a reader-named subfolder."""
        reader_dir = tmp_path / "cn_fund"
        reader_dir.mkdir()
        p = reader_dir / "funding_transactions.xlsx"
        p.write_bytes(b"fake")
        return p

    def _make_result(self, new_transactions=0, new_holdings=0):
        """Build a minimal ProcessResult-like object."""
        class _R:
            pass
        r = _R()
        r.new_transactions = new_transactions
        r.new_holdings = new_holdings
        return r

    def test_upload_called_when_new_rows_and_env_set(
        self, fake_workbook, monkeypatch, tmp_path
    ):
        """new rows > 0 + both env vars set + file under cloud dir → upload called once."""
        from unittest.mock import patch

        cloud_dir = str(fake_workbook.parent.parent)  # tmp_path
        reader_name = fake_workbook.parent.name        # "cn_fund"

        monkeypatch.setenv("UIS_GCS_BUCKET", "test-bucket")
        monkeypatch.setenv("UIS_FINANCE_DIR", cloud_dir)

        mock_result = self._make_result(new_transactions=2, new_holdings=0)

        with patch(
            "src.sources.cn_fund_raw_processor.process_all", return_value=mock_result
        ), patch("src.storage.gcs.upload_source_to_gcs") as mock_upload:
            # Re-import to use patched modules
            import importlib
            import src.sources.reader_hooks as rh
            importlib.reload(rh)  # ensure lazy imports re-resolve
            rh.cn_fund_raw_process(str(fake_workbook), {})

            mock_upload.assert_called_once_with(
                "test-bucket", reader_name, str(fake_workbook)
            )

    def test_upload_not_called_when_no_new_rows(
        self, fake_workbook, monkeypatch, tmp_path
    ):
        """new_transactions=0 and new_holdings=0 → upload NOT called."""
        from unittest.mock import patch

        monkeypatch.setenv("UIS_GCS_BUCKET", "test-bucket")
        monkeypatch.setenv("UIS_FINANCE_DIR", str(fake_workbook.parent.parent))

        mock_result = self._make_result(new_transactions=0, new_holdings=0)

        with patch(
            "src.sources.cn_fund_raw_processor.process_all", return_value=mock_result
        ), patch("src.storage.gcs.upload_source_to_gcs") as mock_upload:
            from src.sources.reader_hooks import cn_fund_raw_process
            cn_fund_raw_process(str(fake_workbook), {})
            mock_upload.assert_not_called()

    def test_upload_not_called_when_env_vars_unset(
        self, fake_workbook, monkeypatch
    ):
        """Neither UIS_GCS_BUCKET nor UIS_FINANCE_DIR set → upload NOT called."""
        from unittest.mock import patch

        monkeypatch.delenv("UIS_GCS_BUCKET", raising=False)
        monkeypatch.delenv("UIS_FINANCE_DIR", raising=False)

        mock_result = self._make_result(new_transactions=1, new_holdings=0)

        with patch(
            "src.sources.cn_fund_raw_processor.process_all", return_value=mock_result
        ), patch("src.storage.gcs.upload_source_to_gcs") as mock_upload:
            from src.sources.reader_hooks import cn_fund_raw_process
            cn_fund_raw_process(str(fake_workbook), {})
            mock_upload.assert_not_called()

    def test_upload_not_called_when_file_not_under_finance_dir(
        self, tmp_path, monkeypatch
    ):
        """File lives outside UIS_FINANCE_DIR → upload NOT called (local dev with env vars)."""
        from unittest.mock import patch

        # finance_dir is a DIFFERENT tmp subdir from where the file lives
        finance_dir = tmp_path / "cloud"
        finance_dir.mkdir()
        file_dir = tmp_path / "local" / "cn_fund"
        file_dir.mkdir(parents=True)
        fake_file = file_dir / "funding_transactions.xlsx"
        fake_file.write_bytes(b"fake")

        monkeypatch.setenv("UIS_GCS_BUCKET", "test-bucket")
        monkeypatch.setenv("UIS_FINANCE_DIR", str(finance_dir))

        mock_result = self._make_result(new_transactions=1, new_holdings=0)

        with patch(
            "src.sources.cn_fund_raw_processor.process_all", return_value=mock_result
        ), patch("src.storage.gcs.upload_source_to_gcs") as mock_upload:
            from src.sources.reader_hooks import cn_fund_raw_process
            cn_fund_raw_process(str(fake_file), {})
            mock_upload.assert_not_called()

    def test_upload_failure_is_non_blocking(
        self, fake_workbook, monkeypatch, caplog
    ):
        """upload_source_to_gcs raises → hook does not raise, warning is logged."""
        import logging
        from unittest.mock import patch

        monkeypatch.setenv("UIS_GCS_BUCKET", "test-bucket")
        monkeypatch.setenv("UIS_FINANCE_DIR", str(fake_workbook.parent.parent))

        mock_result = self._make_result(new_transactions=1, new_holdings=0)

        with patch(
            "src.sources.cn_fund_raw_processor.process_all", return_value=mock_result
        ), patch(
            "src.storage.gcs.upload_source_to_gcs",
            side_effect=RuntimeError("GCS unavailable"),
        ), caplog.at_level(logging.WARNING, logger="src.sources.reader_hooks"):
            from src.sources.reader_hooks import cn_fund_raw_process
            # Must not raise
            cn_fund_raw_process(str(fake_workbook), {})

        assert any(
            "GCS write-back failed" in r.message for r in caplog.records
        ), f"Expected GCS write-back warning in logs; got: {[r.message for r in caplog.records]}"


class TestTypeMapCoverageV717:
    """V7.1.7 — 卖基金/买基金 and 超级转换份额调减 must be recognized end to end."""

    def test_raw_processor_keeps_bank_labels_verbatim(self):
        """卖基金/买基金/超级转换份额调减 pass through unchanged in the processed tab
        (identity) so raw-processor dedup is not broken; the reader hook resolves
        them to transaction types. Mapping them to 赎回/申购 here would duplicate."""
        from src.sources.cn_fund_raw_processor import _map_type
        assert _map_type("卖基金") == "卖基金"
        assert _map_type("买基金") == "买基金"
        assert _map_type("超级转换份额调减") == "超级转换份额调减"

    def test_reader_hook_maps_to_transaction_types(self):
        from src.sources.reader_hooks import _CN_FUND_TYPE_MAP
        assert _CN_FUND_TYPE_MAP["卖基金"] == "sell"
        assert _CN_FUND_TYPE_MAP["买基金"] == "buy"
        assert _CN_FUND_TYPE_MAP["超级转换份额调减"] == "transfer_out"
        # mirror still intact
        assert _CN_FUND_TYPE_MAP["超级转换份额调增"] == "transfer_in"
