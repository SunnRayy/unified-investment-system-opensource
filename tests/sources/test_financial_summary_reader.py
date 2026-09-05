"""Config-engine smoke tests for Financial Summary reader (B5 — legacy FinancialSummaryReader deleted).

The legacy FinancialSummaryReader was deleted in Workstream B5.
sync_financial_summary() now reads the Excel sheets directly with pandas.
These tests verify it reads the real fixture correctly.
"""
import pytest
import openpyxl
from datetime import datetime
from pathlib import Path

pytestmark = pytest.mark.pipeline

FIXTURE_DIR = Path(__file__).parent.parent / "fixtures" / "readers"
FS_FIXTURE = FIXTURE_DIR / "Financial_Summary_new.xlsx"


class TestFinancialSummarySync:
    def _make_config(self, data_dir: str, workbook_name: str) -> dict:
        return {
            "source_registry": {
                "financial_summary": {
                    "enabled": True,
                    "data_dir": str(data_dir),
                    "file_patterns": {"workbook": workbook_name},
                }
            }
        }

    def test_read_skips_header_rows(self, tmp_path):
        """sync_financial_summary reads both sheets, skipping 3 header rows."""
        from src.sync.financial_summary_sync import sync_financial_summary

        wb = openpyxl.Workbook()
        ws_bs = wb.active
        ws_bs.title = "资产负债"
        ws_bs.append(["", "", ""])
        ws_bs.append(["", "Report", ""])
        ws_bs.append(["", "", ""])
        ws_bs.append(["Date", "Total Assets", "Total Liabilities", "Net Worth"])
        ws_bs.append([datetime(2025, 1, 1), 100000, 20000, 80000])
        ws_bs.append([datetime(2025, 2, 1), 105000, 19000, 86000])

        ws_ie = wb.create_sheet("月度收支")
        ws_ie.append(["", "", ""])
        ws_ie.append(["", "Income Statement", ""])
        ws_ie.append(["", "", ""])
        ws_ie.append(["Month", "Income", "Expense", "Savings"])
        ws_ie.append([datetime(2025, 1, 1), 5000, 2000, 3000])

        path = tmp_path / "Financial Summary_new.xlsx"
        wb.save(path)

        config = self._make_config(str(tmp_path), "Financial Summary_new.xlsx")
        result = sync_financial_summary(config)

        assert not result["balance_sheet"].empty
        assert "Total Assets" in result["balance_sheet"].columns
        assert result["balance_sheet"]["Total Assets"].iloc[0] == 100000

        assert not result["income_expense"].empty
        assert "Income" in result["income_expense"].columns
        assert result["income_expense"]["Income"].iloc[0] == 5000

    def test_file_not_found(self, tmp_path):
        from src.sync.financial_summary_sync import sync_financial_summary
        config = self._make_config(str(tmp_path), "nonexistent.xlsx")
        result = sync_financial_summary(config)
        assert result["balance_sheet"].empty
        assert result["income_expense"].empty

    def test_disabled_returns_empty(self):
        from src.sync.financial_summary_sync import sync_financial_summary
        config = {"source_registry": {"financial_summary": {"enabled": False}}}
        result = sync_financial_summary(config)
        assert result["holdings"].empty
        assert result["transactions"].empty

    def test_broken_format_handled(self, tmp_path):
        """File with fewer than 4 rows should not crash."""
        from src.sync.financial_summary_sync import sync_financial_summary

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "资产负债"
        ws.append(["Row 1"])
        path = tmp_path / "Financial Summary_new.xlsx"
        wb.save(path)

        config = self._make_config(str(tmp_path), "Financial Summary_new.xlsx")
        result = sync_financial_summary(config)
        # Should return empty but not crash
        assert result["balance_sheet"].empty

    def test_asset_ids_added(self, tmp_path):
        """balance_sheet rows must have synthetic asset_id starting with BS_."""
        from src.sync.financial_summary_sync import sync_financial_summary

        wb = openpyxl.Workbook()
        ws_bs = wb.active
        ws_bs.title = "资产负债"
        for _ in range(3):
            ws_bs.append([""])
        ws_bs.append(["Date", "Total Assets"])
        ws_bs.append([datetime(2025, 1, 1), 100000])

        wb.create_sheet("月度收支")
        path = tmp_path / "Financial Summary_new.xlsx"
        wb.save(path)

        config = self._make_config(str(tmp_path), "Financial Summary_new.xlsx")
        result = sync_financial_summary(config)
        assert "asset_id" in result["balance_sheet"].columns
        assert result["balance_sheet"]["asset_id"].iloc[0].startswith("BS_")
