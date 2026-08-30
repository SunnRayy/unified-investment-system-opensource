"""CN Fund raw transaction/holdings processor.

Reads raw_transactions_paste and raw_holdings_paste sheets from the CN Fund Excel
workbook, finds rows not yet in the organized sheets (基金交易记录 / 基金持仓汇总),
and appends them.

This runs before CNFundReader so that the organized sheets are always up to date.
"""
import logging
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.utils.datetime import from_excel as _xl_to_datetime

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Type mapping tables
# ---------------------------------------------------------------------------

TYPE_MAP = {
    # Buy types → 申购
    '申购': '申购', '买入': '申购', '定投': '申购',
    # Sell types → 赎回
    '赎回': '赎回', '卖出': '赎回', '普通取现': '赎回',
    '活期宝转出': '赎回', '预约取现': '赎回',
    '信用卡还款': '赎回', '解散': '赎回',
    # Money market
    '快速取现': '快速取现',
    '充值': '活期宝即充即用', '自动充值': '活期宝即充即用',
    '活期宝即充即用': '活期宝即充即用',
    # Dividends (passthrough)
    '现金分红': '现金分红', '分红': '现金分红',
    '红利再投资': '红利再投资',
    '结转收益': '红利再投资', '每日收益': '红利再投资',
    # Conversions
    '超级转换-转入': '超级转换-转入',
    '超级转换-转出': '超级转换-转出',
    '超级转换份额调增': '超级转换份额调增',
    '转换': '转换', '互转': '转换', '调仓': '转换', '组合调仓': '转换',
    # Newer bank labels — kept as identity so the processed tab mirrors the bank
    # verbatim (the Huinsight reader hook resolves them to sell/buy/transfer_out).
    # Mapping these to 赎回/申购 would break raw-processor dedup against rows
    # already written as 卖基金/超级转换份额调减 → duplicate processed rows.
    '卖基金': '卖基金', '买基金': '买基金',
    '超级转换份额调减': '超级转换份额调减',
}

REASON_MAP = {
    '申购': '手动买入', '买入': '手动买入',
    '定投': '定投',
    '赎回': '手动卖出', '卖出': '手动卖出',
    '普通取现': '普通取现', '快速取现': '快速取现',
    '活期宝转出': '活期宝转出', '预约取现': '预约取现',
    '信用卡还款': '信用卡还款', '解散': '解散',
    '充值': '活期宝充值', '自动充值': '活期宝充值', '活期宝即充即用': '活期宝充值',
    '现金分红': 'Cash Dividend', '分红': 'Cash Dividend',
    '红利再投资': 'Dividend Reinvestment',
    '结转收益': '结转收益', '每日收益': '每日收益',
}


# ---------------------------------------------------------------------------
# Result dataclasses
# ---------------------------------------------------------------------------

@dataclass
class TransactionProcessResult:
    new_count: int  # Number of new rows appended to 基金交易记录


@dataclass
class HoldingsProcessResult:
    new_count: int  # Number of new snapshot rows appended to 基金持仓汇总


@dataclass
class ProcessAllResult:
    new_transactions: int
    new_holdings: int


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _map_type(raw_type: str) -> str:
    if raw_type in TYPE_MAP:
        return TYPE_MAP[raw_type]
    logger.warning(f"CN Fund raw processor: unknown type '{raw_type}', using as-is")
    return raw_type


def _map_reason(raw_type: str) -> str:
    return REASON_MAP.get(raw_type, raw_type)


def _amount_key(v) -> Optional[float]:
    if v is None:
        return None
    try:
        return round(float(v), 2)
    except (TypeError, ValueError):
        return None


def _to_datetime(v) -> Optional[datetime]:
    """Normalize a value to a datetime, handling Excel float serials."""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v
    if hasattr(v, 'year'):  # date object
        return datetime(v.year, v.month, v.day)
    if isinstance(v, (int, float)):
        try:
            return _xl_to_datetime(int(v))
        except Exception:
            return None
    return None


def _to_date(v):
    """Extract a date object from a datetime, date, or Excel float serial."""
    dt = _to_datetime(v)
    return dt.date() if dt is not None else None


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def process_raw_transactions(workbook_path: Path) -> TransactionProcessResult:
    """Append new rows from raw_transactions_paste to 基金交易记录.

    RAW columns:  确认日期, 基金代码, 基金简称, 业务类型, 确认状态, 确认份额, 确认金额, 手续费, 确认净值, 关联银行卡
    OUT columns:  交易日期, 基金代码, 基金名称, 操作类型, 交易金额, 交易份额, 交易时基金单位净值, 手续费, 交易原因
    """
    wb = openpyxl.load_workbook(workbook_path)
    raw_ws = wb['raw_transactions_paste']
    processed_ws = wb['基金交易记录']

    # Build dedup set from existing processed rows (skip header row 1)
    # Processed cols: [0]=交易日期 [1]=基金代码 [3]=操作类型 [4]=交易金额
    existing_keys: set = set()
    for row in processed_ws.iter_rows(min_row=2, values_only=True):
        tx_date = _to_date(row[0])
        fund_code = str(row[1]) if row[1] is not None else None
        op_type = str(row[3]) if row[3] is not None else None
        amount = _amount_key(row[4])
        if tx_date is not None:
            existing_keys.add((str(tx_date), fund_code, op_type, amount))

    # Raw cols: [0]=确认日期 [1]=基金代码 [2]=基金简称 [3]=业务类型 [4]=确认状态
    #           [5]=确认份额 [6]=确认金额 [7]=手续费 [8]=确认净值 [9]=关联银行卡
    new_count = 0
    for row in raw_ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue

        confirm_date = row[0]
        fund_code = str(row[1]) if row[1] is not None else None
        fund_name = row[2]
        raw_type = str(row[3]) if row[3] is not None else ''
        status = str(row[4]) if row[4] is not None else ''
        shares = row[5]
        raw_amount = row[6]
        fee = row[7]
        nav = row[8]

        # Skip formula strings (e.g. '=F4*I4') — compute from shares × nav instead
        if isinstance(raw_amount, str) and raw_amount.startswith('='):
            try:
                amount = round(float(shares) * float(nav), 4) if shares is not None and nav is not None else None
            except (TypeError, ValueError):
                amount = None
        else:
            amount = raw_amount

        # Only process confirmed rows
        if status != '成功':
            continue

        tx_date = _to_date(confirm_date)
        mapped_type = _map_type(raw_type)
        key = (str(tx_date), fund_code, mapped_type, _amount_key(amount))

        if key in existing_keys:
            continue

        shares_val = float(shares) if shares is not None else 0.0
        processed_ws.append([
            confirm_date,           # 交易日期
            fund_code,              # 基金代码
            fund_name,              # 基金名称
            mapped_type,            # 操作类型
            amount,                 # 交易金额
            shares_val,             # 交易份额
            nav,                    # 交易时基金单位净值
            fee,                    # 手续费
            _map_reason(raw_type),  # 交易原因
        ])
        existing_keys.add(key)
        new_count += 1

    if new_count > 0:
        wb.save(workbook_path)

    return TransactionProcessResult(new_count=new_count)


def process_raw_holdings(workbook_path: Path) -> HoldingsProcessResult:
    """Aggregate and append new snapshots from raw_holdings_paste to 基金持仓汇总.

    RAW columns:  基金代码, 基金简称, 基金类型, 净值日期, 单位净值, 持有份额, 参考市值, 关联银行卡
    OUT columns:  Asset_ID, Asset_Name, Asset_Type_Raw, Snapshot_Date, Market_Price_Unit, Quantity, Market_Value_Raw
    """
    wb = openpyxl.load_workbook(workbook_path)
    raw_ws = wb['raw_holdings_paste']
    processed_ws = wb['基金持仓汇总']

    # Aggregate raw rows by (fund_code, nav_date).
    #
    # Column layout detection: the real bank export has 8 columns matching the header
    # exactly (基金代码 at col 0).  Test fixtures prepend an extra "record date" datetime
    # making 9 columns (基金代码 at col 1).  We detect the format per-row by checking
    # whether col 0 is a datetime (test fixture) or a string (real file).
    aggregated: dict = {}
    for row in raw_ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue

        # Detect offset: if col 0 is a datetime it's the test-fixture 9-column format
        offset = 1 if isinstance(row[0], datetime) else 0

        if row[offset] is None:
            continue

        fund_code = str(row[offset])
        fund_name = row[offset + 1] if len(row) > offset + 1 else None
        fund_type = row[offset + 2] if len(row) > offset + 2 else None
        nav_date  = row[offset + 3] if len(row) > offset + 3 else None
        nav       = row[offset + 4] if len(row) > offset + 4 else None
        shares_raw = row[offset + 5] if len(row) > offset + 5 else None
        shares    = float(shares_raw) if shares_raw is not None else 0.0

        # 参考市值 at offset+6; may be 关联银行卡 (string) if the column is absent.
        # Fall back to shares × nav in that case.
        market_raw = row[offset + 6] if len(row) > offset + 6 else None
        try:
            market_value = float(market_raw) if market_raw is not None else shares * (float(nav) if nav else 0.0)
        except (TypeError, ValueError):
            market_value = shares * (float(nav) if nav is not None else 0.0)

        nav_date_dt = _to_datetime(nav_date)
        if nav_date_dt is None:
            logger.warning(f"CN Fund raw processor: skipping holdings row with unparseable nav_date={nav_date!r}")
            continue
        agg_key = (fund_code, nav_date_dt.date())
        if agg_key not in aggregated:
            aggregated[agg_key] = {
                'Asset_ID': fund_code,
                'Asset_Name': fund_name,
                'Asset_Type_Raw': fund_type,
                'Snapshot_Date': nav_date_dt,  # always a datetime
                'Market_Price_Unit': nav,
                'Quantity': 0.0,
                'Market_Value_Raw': 0.0,
            }
        aggregated[agg_key]['Quantity'] += shares
        aggregated[agg_key]['Market_Value_Raw'] += market_value

    # Build dedup set from existing processed rows (skip header)
    # Processed cols: [0]=Asset_ID [3]=Snapshot_Date
    existing_keys: set = set()
    for row in processed_ws.iter_rows(min_row=2, values_only=True):
        asset_id = str(row[0]) if row[0] is not None else None
        snap_date = _to_date(row[3])
        if asset_id is not None and snap_date is not None:
            existing_keys.add((asset_id, snap_date))

    new_count = 0
    for (fund_code, nav_date_d), data in aggregated.items():
        if (fund_code, nav_date_d) in existing_keys:  # nav_date_d is already a date object
            continue

        processed_ws.append([
            data['Asset_ID'],
            data['Asset_Name'],
            data['Asset_Type_Raw'],
            data['Snapshot_Date'],
            data['Market_Price_Unit'],
            data['Quantity'],
            data['Market_Value_Raw'],
        ])
        existing_keys.add((fund_code, nav_date_d))
        new_count += 1

    if new_count > 0:
        wb.save(workbook_path)

    return HoldingsProcessResult(new_count=new_count)


def process_all(workbook_path: Path) -> ProcessAllResult:
    """Process both raw transactions and holdings. Entry point for cn_fund_sync."""
    tx = process_raw_transactions(workbook_path)
    hld = process_raw_holdings(workbook_path)
    return ProcessAllResult(new_transactions=tx.new_count, new_holdings=hld.new_count)
