"""Reference-sheet exporter — D1 of the FS Reference-Data automation.

Generates `UIS_Reference_Data.xlsx` (a single `Reference` sheet with a stable
row-per-label layout) so the owner can external-link their Financial Summary
Excel into it instead of entering values by hand every month.

Layout (stable — do NOT reorder columns; external links depend on position):
    A: Label
    B: Value_CNY
    C: Value_USD
    D: Qty
    E: Price
    F: AsOf

Entry point: export_reference_sheet(connector, config, out_dir=None) -> Path

Safety guarantee:
    The output file is ALWAYS named `UIS_Reference_Data.xlsx`.
    The function refuses to write to any other filename and refuses to overwrite
    any known source workbook (Financial Summary_new.xlsx, funding_transactions.xlsx,
    etc.). Violations raise ValueError BEFORE any file I/O.

Non-fatal policy:
    If a query or price look-up fails, the affected cell is left blank and a
    warning is logged. The sync never fails because of a reference-sheet error.
"""

from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import yaml

from src.config import _resolve_config_file
from src.data_manager.currency_converter import get_currency_service

logger = logging.getLogger(__name__)

# ── Constants ──────────────────────────────────────────────────────────────────

OUTPUT_FILENAME = "UIS_Reference_Data.xlsx"
SHEET_NAME = "Reference"

# ── WS-2 (Attribution & Flows program): 月度快照 (monthly snapshot) sheet ──────
# Owner decisions (2026-07-19, docs/plans/2026-07-19-attribution-flows-program.md §WS-2):
#   - Backfill starts 2026-01 (no earlier history is reconstructed).
#   - Append-only: once a calendar month is CLOSED (strictly before the export's current
#     month), its (month, asset_id) rows are frozen forever — a re-export must NEVER
#     rewrite them, even if the DB's historical values later change. The owner's Excel
#     VLOOKUPs external-link into this sheet by month; rewriting history would silently
#     break every prior month's formula results.
#   - The CURRENT (still-open) month is the one exception: it is recomputed and replaced
#     on every export until the month closes, so the owner sees a live running total.
MONTHLY_SHEET_NAME = "月度快照"
_MONTHLY_BACKFILL_START_MONTH = "2026-01"
_MONTHLY_HEADER_ROW = ["Month", "Asset_ID", "Asset_Name", "Source", "Qty", "Market_Value_CNY"]

# Source workbooks that must NEVER be overwritten.
# Names only (case-insensitive match against the resolved target path stem).
_FORBIDDEN_FILENAMES: frozenset[str] = frozenset(
    name.lower() for name in [
        "financial summary_new",
        "funding_transactions",
        "gold_transactions",
        "rsu_transactions",
        "insurance",
    ]
)

_DEFAULT_CONFIG_PATH = Path("config/reference_sheet.yaml")

# Column positions (0-based) in the Reference sheet
_COL_LABEL = 0
_COL_CNY = 1
_COL_USD = 2
_COL_QTY = 3
_COL_PRICE = 4
_COL_ASOF = 5

_HEADER_ROW = ["Label", "Value_CNY", "Value_USD", "Qty", "Price", "AsOf"]


# ── Config loader ──────────────────────────────────────────────────────────────

def _load_reference_config(
    config_path: Optional[Path] = None,
) -> List[Dict[str, Any]]:
    """Load config/reference_sheet.yaml and return the `items` list.

    Falls back to the committed reference_sheet.example.yaml template when
    the real (gitignored) file is absent — see src.config._resolve_config_file.
    """
    path = config_path or _DEFAULT_CONFIG_PATH
    path = _resolve_config_file(Path(path))
    with path.open("r", encoding="utf-8") as fh:
        data = yaml.safe_load(fh)
    items = data.get("items", [])
    if not items:
        logger.warning("reference_sheet.yaml has no items; output will be empty")
    return items


# ── Safety guard ───────────────────────────────────────────────────────────────

def _assert_safe_output_path(out_path: Path) -> None:
    """Raise ValueError if `out_path` is unsafe (wrong name or known source workbook)."""
    if out_path.name != OUTPUT_FILENAME:
        raise ValueError(
            f"Safety guard: output filename must be '{OUTPUT_FILENAME}', "
            f"got '{out_path.name}'"
        )
    stem_lower = out_path.stem.lower()
    if stem_lower in _FORBIDDEN_FILENAMES:
        raise ValueError(
            f"Safety guard: refusing to write to protected source workbook path: {out_path}"
        )
    # Also reject if ANY parent directory component contains a forbidden name
    # (belt-and-suspenders; guards against crafted paths like
    # /Finance/Financial Summary_new.xlsx/UIS_Reference_Data.xlsx)
    for part in out_path.parts[:-1]:
        if part.lower().replace(".xlsx", "") in _FORBIDDEN_FILENAMES:
            raise ValueError(
                f"Safety guard: suspicious path component '{part}' near source workbook: {out_path}"
            )


# ── DB query helpers ──────────────────────────────────────────────────────────

def _coauthority_sources() -> "frozenset[str]":
    """Return AuthorityResolver's co-authority source set (e.g. {Schwab_CSV, Broker_IBKR}).

    Wrapped in its own function (rather than imported at module scope) so tests can patch
    `src.sync.reference_export.AuthorityResolver` directly. Failures degrade to an empty
    set — i.e. no co-authority share recovery — rather than breaking the whole export
    (matches the module's non-fatal policy).
    """
    from src.identity.authority_resolver import AuthorityResolver

    return AuthorityResolver().coauthority_sources()


def _query_broker_consolidated_share(
    connector: Any,
    source_system: str,
) -> Tuple[Optional[float], Optional[float], Optional[str]]:
    """Return (cny_share, usd_share, as_of) for `source_system`'s slice of holdings that
    are now folded into a `Consolidated` row (C3.4 co-authority merge).

    Background (WS-2, Attribution & Flows program): `_consolidate_coauthority_holdings`
    (src/sync/phases/_shadow.py) merges co-authority broker rows — currently Schwab_CSV +
    Broker_IBKR — for the same asset into a single `source_system='Consolidated'` holdings
    row and marks the contributing broker rows `is_shadow=TRUE` so downstream `is_shadow=
    FALSE` queries don't double-count. `_query_source_sum` below filters `is_shadow=FALSE`,
    so a co-authority broker's reported total silently DROPS every asset it shares with the
    other broker (e.g. IBKR's SGOV/VOO/IEF share) — this is exactly why the owner has been
    hand-patching the Financial Summary import every month.

    This helper recovers that share: `source_system`'s own latest (asset_id, source_system)
    row — regardless of is_shadow — restricted to asset_ids that currently have an ACTIVE
    Consolidated row (so a fully-liquidated / de-consolidated asset is correctly excluded,
    and a still-shadowed-but-stale duplicate never leaks in). This reuses the exact same
    asset relationship the `consolidated_equals_sum` integrity check (#15) verifies — see
    `src/validation/data_integrity_gate.py::_check_consolidated_equals_sum` (broker_latest /
    broker_sums CTEs) — just scoped to one broker instead of summed across all of them.

    Per-(asset, source_system) MAX(snapshot_date) — never a global MAX (Rule 3).
    """
    try:
        row = connector.execute(
            """
            WITH latest_broker AS (
                SELECT asset_id, MAX(snapshot_date) AS max_date
                FROM holdings
                WHERE source_system = ?
                GROUP BY asset_id
            ),
            broker_rows AS (
                SELECT h.asset_id, h.market_value, h.quantity, h.market_price_unit,
                       h.currency, h.snapshot_date
                FROM holdings h
                JOIN latest_broker lb
                  ON h.asset_id      = lb.asset_id
                 AND h.snapshot_date = lb.max_date
                WHERE h.source_system = ?
            ),
            active_consolidated AS (
                SELECT DISTINCT asset_id
                FROM holdings
                WHERE source_system = 'Consolidated'
                  AND is_shadow = FALSE
            )
            SELECT
                SUM(br.market_value)                                    AS cny_total,
                SUM(CASE WHEN br.currency = 'USD'
                         THEN br.quantity * br.market_price_unit
                         ELSE NULL END)                                 AS usd_total,
                MAX(br.snapshot_date)                                   AS as_of
            FROM broker_rows br
            JOIN active_consolidated ac ON ac.asset_id = br.asset_id
            """,
            (source_system, source_system),
        ).fetchone()
    except Exception as exc:
        logger.warning("broker_consolidated_share query failed for %s: %s", source_system, exc)
        return None, None, None

    if row is None or row[0] is None:
        return None, None, None

    cny = float(row[0]) if row[0] is not None else None
    usd = float(row[1]) if row[1] is not None else None
    as_of = str(row[2]) if row[2] is not None else None
    return cny, usd, as_of


def _query_source_sum(
    connector: Any,
    source_systems: List[str],
) -> Tuple[Optional[float], Optional[float], Optional[str]]:
    """Return (cny_total, usd_total, as_of_iso) for the given source_systems.

    cny_total: SUM(market_value) in CNY (market_value is always CNY in Huinsight)
    usd_total: SUM(quantity * market_price_unit) for USD-currency holdings
    as_of_iso: latest snapshot_date across the queried rows (ISO string)

    Uses per-asset MAX(snapshot_date) — never a global MAX (Rule 3).

    Co-authority note (WS-2): for any `source_system` in `source_systems` that is a
    co-authority broker (Schwab_CSV, Broker_IBKR), this ALSO folds in that broker's share
    of assets now merged into `Consolidated` holdings rows — see
    `_query_broker_consolidated_share`. Without this, a co-authority broker's total
    silently excludes every asset it shares with the other broker.
    """
    if not source_systems:
        return None, None, None

    systems_placeholder = ", ".join("?" * len(source_systems))
    try:
        row = connector.execute(
            f"""
            WITH latest AS (
                SELECT asset_id, source_system,
                       MAX(snapshot_date) AS snap_date
                FROM holdings
                WHERE source_system IN ({systems_placeholder})
                  AND is_shadow = FALSE
                GROUP BY asset_id, source_system
            )
            SELECT
                SUM(h.market_value)                                    AS cny_total,
                SUM(CASE WHEN h.currency = 'USD'
                         THEN h.quantity * h.market_price_unit
                         ELSE NULL END)                                AS usd_total,
                MAX(l.snap_date)                                       AS as_of
            FROM holdings h
            JOIN latest l
              ON h.asset_id      = l.asset_id
             AND h.source_system = l.source_system
             AND h.snapshot_date = l.snap_date
            WHERE h.is_shadow = FALSE
            """,
            source_systems,
        ).fetchone()
    except Exception as exc:
        logger.warning("source_sum query failed for %s: %s", source_systems, exc)
        return None, None, None

    if row is None:
        return None, None, None

    cny = float(row[0]) if row[0] is not None else None
    usd = float(row[1]) if row[1] is not None else None
    as_of = str(row[2]) if row[2] is not None else None

    # ── Co-authority share recovery (C3.4 / WS-2) ──────────────────────────
    try:
        coauth_sources = _coauthority_sources()
    except Exception as exc:
        logger.warning("coauthority_sources() lookup failed: %s", exc)
        coauth_sources = frozenset()

    for src in source_systems:
        if src not in coauth_sources:
            continue
        share_cny, share_usd, share_as_of = _query_broker_consolidated_share(connector, src)
        if share_cny is not None:
            cny = (cny or 0.0) + share_cny
        if share_usd is not None:
            usd = (usd or 0.0) + share_usd
        if share_as_of is not None and (as_of is None or share_as_of > as_of):
            as_of = share_as_of

    return cny, usd, as_of


def _query_asset(
    connector: Any,
    asset_id: str,
) -> Tuple[Optional[float], Optional[float], Optional[float], Optional[float], Optional[str]]:
    """Return (cny, usd, qty, price, as_of) for a single asset.

    cny  : market_value in CNY
    usd  : quantity * market_price_unit (native price; meaningful only if currency='USD')
    qty  : raw quantity
    price: market_price_unit
    as_of: latest snapshot_date ISO string
    """
    try:
        row = connector.execute(
            """
            SELECT
                h.market_value,
                h.quantity * h.market_price_unit  AS usd_value,
                h.quantity,
                h.market_price_unit,
                h.currency,
                h.snapshot_date
            FROM holdings h
            WHERE h.asset_id = ?
              AND h.is_shadow = FALSE
            ORDER BY h.snapshot_date DESC
            LIMIT 1
            """,
            (asset_id,),
        ).fetchone()
    except Exception as exc:
        logger.warning("asset query failed for %s: %s", asset_id, exc)
        return None, None, None, None, None

    if row is None:
        logger.debug("asset_id '%s' not found in holdings (will emit blank)", asset_id)
        return None, None, None, None, None

    cny = float(row[0]) if row[0] is not None else None
    # usd_value is only meaningful for USD assets
    currency = str(row[4]) if row[4] else "CNY"
    usd = float(row[1]) if (row[1] is not None and currency == "USD") else None
    qty = float(row[2]) if row[2] is not None else None
    price = float(row[3]) if row[3] is not None else None
    as_of = str(row[5]) if row[5] is not None else None
    return cny, usd, qty, price, as_of


def _get_fx_rate(connector: Any) -> Tuple[Optional[float], Optional[str]]:
    """Return live USD/CNY rate from the shared currency service (same source as Dashboard).

    The `connector` parameter is accepted for API compatibility with callers but is not
    used — the rate comes from the live currency service, not from holdings.

    Returns (rate, None) on success; (7.0, None) with a warning log on failure.
    """
    try:
        rate = get_currency_service().get_latest_rate("USD", "CNY")
        if rate is not None:
            logger.debug("USD/CNY rate from currency service: %.4f", rate)
            return float(rate), None
        logger.warning("Currency service returned None for USD/CNY — using fallback 7.0")
        return 7.0, None
    except Exception as exc:
        logger.warning("Currency service error fetching USD/CNY: %s — using fallback 7.0", exc)
        return 7.0, None


def _get_market_price(
    connector: Any,
    ticker: str,
) -> Tuple[Optional[float], Optional[str]]:
    """Return (price, as_of_iso) for a bare ticker from market_daily.

    Uses per-ticker MAX(date) — never a global MAX.
    """
    try:
        row = connector.execute(
            """
            SELECT close, date
            FROM market_daily
            WHERE code = ?
              AND date = (SELECT MAX(date) FROM market_daily WHERE code = ?)
            LIMIT 1
            """,
            (ticker, ticker),
        ).fetchone()
    except Exception as exc:
        logger.warning("market_price query failed for %s: %s", ticker, exc)
        return None, None

    if row is None:
        logger.debug("No market_daily entry for ticker '%s'", ticker)
        return None, None

    price = float(row[0]) if row[0] is not None else None
    as_of = str(row[1]) if row[1] is not None else None
    return price, as_of


# ── 月度快照 (monthly snapshot) — WS-2 ──────────────────────────────────────────

def _compute_monthly_snapshot_rows(
    connector: Any,
    start_month: str,
    end_month_inclusive: str,
) -> List[Tuple[str, str, str, str, Optional[float], Optional[float]]]:
    """Compute month-end (qty, market_value_cny, source) per (month, asset_id).

    For each (month, asset_id) pair, picks the single `is_shadow=FALSE` holdings row
    whose `snapshot_date` is the LATEST one falling within that calendar month — i.e. a
    per-asset, per-month MAX(snapshot_date). This is intentionally never a global
    MAX(snapshot_date) (Rule 3): each asset's own most-recent-within-the-month snapshot is
    used independently, so a QDII fund reported a few days behind a broker's daily feed
    still gets its correct month-end row instead of being silently dropped or misdated.

    `start_month` / `end_month_inclusive`: 'YYYY-MM' strings, inclusive on both ends.

    Returns a list of (month, asset_id, asset_name, source_system, qty, market_value_cny)
    tuples, sorted by (month, asset_id). Non-fatal: query failure logs a warning and
    returns [] (the caller then just preserves whatever pre-existing sheet rows exist).
    """
    start_date = f"{start_month}-01"
    end_year, end_mon = (int(x) for x in end_month_inclusive.split("-"))
    if end_mon == 12:
        end_exclusive = f"{end_year + 1}-01-01"
    else:
        end_exclusive = f"{end_year}-{end_mon + 1:02d}-01"

    try:
        rows = connector.execute(
            """
            WITH monthly AS (
                SELECT
                    strftime(snapshot_date, '%Y-%m') AS month,
                    asset_id,
                    MAX(snapshot_date)                AS month_end_date
                FROM holdings
                WHERE is_shadow = FALSE
                  AND snapshot_date >= ?
                  AND snapshot_date <  ?
                GROUP BY 1, 2
            ),
            ranked AS (
                SELECT
                    m.month, m.asset_id, h.asset_name, h.source_system,
                    h.quantity, h.market_value,
                    ROW_NUMBER() OVER (
                        PARTITION BY m.month, m.asset_id
                        ORDER BY h.source_system
                    ) AS rn
                FROM monthly m
                JOIN holdings h
                  ON h.asset_id      = m.asset_id
                 AND h.snapshot_date = m.month_end_date
                WHERE h.is_shadow = FALSE
            )
            SELECT month, asset_id, asset_name, source_system, quantity, market_value
            FROM ranked
            WHERE rn = 1
            ORDER BY month, asset_id
            """,
            (start_date, end_exclusive),
        ).fetchall()
    except Exception as exc:
        logger.warning("monthly_snapshot query failed: %s", exc)
        return []

    return [
        (
            str(r[0]),
            str(r[1]),
            str(r[2]) if r[2] is not None else "",
            str(r[3]) if r[3] is not None else "",
            float(r[4]) if r[4] is not None else None,
            float(r[5]) if r[5] is not None else None,
        )
        for r in rows
    ]


def _load_existing_monthly_snapshot(out_path: Path) -> Dict[Tuple[str, str], List]:
    """Read pre-existing 月度快照 rows from `out_path`, keyed by (month, asset_id).

    Returns {} if the file doesn't exist yet, has no 月度快照 sheet, or fails to parse —
    non-fatal, since "no prior file" just means there is no history to freeze yet (first
    export of the program).
    """
    if not out_path.exists():
        return {}
    try:
        import openpyxl

        wb = openpyxl.load_workbook(str(out_path), data_only=True)
    except Exception as exc:
        logger.warning(
            "Could not open existing reference workbook for 月度快照 merge (%s) — "
            "treating as no prior history.", exc,
        )
        return {}

    if MONTHLY_SHEET_NAME not in wb.sheetnames:
        return {}

    ws = wb[MONTHLY_SHEET_NAME]
    existing: Dict[Tuple[str, str], List] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None or row[0] is None or row[1] is None:
            continue
        month, asset_id = str(row[0]), str(row[1])
        existing[(month, asset_id)] = list(row)
    return existing


def _merge_monthly_snapshot(
    existing: Dict[Tuple[str, str], List],
    computed: List[Tuple],
    current_month: str,
) -> List[List]:
    """Merge freshly computed 月度快照 rows with the pre-existing sheet — append-only.

    Owner decision (2026-07-19): once a month is CLOSED (`month < current_month`), its
    rows are frozen forever — an existing (month, asset_id) row is NEVER overwritten by a
    re-export, even if the DB's historical values would now compute differently (e.g. a
    late correction). The owner's Excel VLOOKUPs external-link into this sheet by month;
    rewriting a closed month would silently change every prior period's formula results.

    Only genuinely NEW (month, asset_id) keys for closed months — first-time backfill, or
    a month that has just closed since the last export — are added from `computed`.

    The CURRENT (still-open) month is the one exception: it is fully replaced by the
    freshly computed rows on every run (including dropping any pre-existing current-month
    row that `computed` no longer reproduces, e.g. a fully-liquidated asset) so the
    owner's running total stays live until the month closes.
    """
    computed_by_key: Dict[Tuple[str, str], List] = {(r[0], r[1]): list(r) for r in computed}

    merged: Dict[Tuple[str, str], List] = {}

    # 1. Frozen (closed-month) rows: existing wins verbatim — never rewritten.
    for key, row in existing.items():
        if key[0] < current_month:
            merged[key] = row

    # 2. Newly-closed / first-time-backfilled months: fill gaps from computed only.
    for key, row in computed_by_key.items():
        if key[0] < current_month and key not in merged:
            merged[key] = row

    # 3. Current (open) month: always fresh from `computed`.
    for key, row in computed_by_key.items():
        if key[0] == current_month:
            merged[key] = row

    return [merged[k] for k in sorted(merged.keys())]


# ── Row builder ───────────────────────────────────────────────────────────────

def _build_rows(
    connector: Any,
    items: List[Dict[str, Any]],
) -> List[List]:
    """Evaluate each config item against the DB and return a list of data rows.

    Each row: [label, cny, usd, qty, price, as_of]
    """
    rows: List[List] = []
    fx_rate, fx_as_of = None, None  # lazily computed once

    for item in items:
        label = item.get("fs_label", "")
        source = item.get("source", {})
        emit = item.get("emit", {})

        cny_val: Optional[float] = None
        usd_val: Optional[float] = None
        qty_val: Optional[float] = None
        price_val: Optional[float] = None
        as_of_val: Optional[str] = None

        stype = source.get("type")

        if stype == "source_sum":
            src_systems = source.get("source_systems", [])
            try:
                cny_sum, usd_sum, as_of = _query_source_sum(connector, src_systems)
                if emit.get("cny"):
                    cny_val = cny_sum
                if emit.get("usd"):
                    usd_val = usd_sum
                as_of_val = as_of
            except Exception as exc:
                logger.warning("source_sum eval failed for '%s': %s", label, exc)

        elif stype == "asset":
            asset_id = source.get("asset_id", "")
            try:
                cny, usd, qty, price, as_of = _query_asset(connector, asset_id)
                if emit.get("cny"):
                    cny_val = cny
                if emit.get("usd"):
                    usd_val = usd
                if emit.get("qty"):
                    qty_val = qty
                if emit.get("price"):
                    price_val = price
                as_of_val = as_of
            except Exception as exc:
                logger.warning("asset eval failed for '%s' (%s): %s", label, asset_id, exc)

        elif stype == "price_row":
            price_type = source.get("price_type", "")
            if price_type == "fx_usdcny":
                if fx_rate is None:
                    fx_rate, fx_as_of = _get_fx_rate(connector)
                if emit.get("fx"):
                    price_val = fx_rate
                as_of_val = fx_as_of
            elif price_type == "market_price":
                ticker = source.get("ticker", "")
                try:
                    p, as_of = _get_market_price(connector, ticker)
                    if emit.get("price"):
                        price_val = p
                    as_of_val = as_of
                except Exception as exc:
                    logger.warning("market_price eval failed for '%s' (%s): %s", label, ticker, exc)
        else:
            logger.warning("Unknown source type '%s' for item '%s'", stype, label)

        rows.append([label, cny_val, usd_val, qty_val, price_val, as_of_val])

    return rows


# ── Excel writer ──────────────────────────────────────────────────────────────

def _write_xlsx(
    out_path: Path,
    rows: List[List],
    generated_at: datetime,
    monthly_rows: Optional[List[List]] = None,
) -> None:
    """Write the reference sheet (+ 月度快照 sheet, WS-2) to an Excel file using openpyxl."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    # Row 1: "As of <timestamp>" banner
    ws.cell(row=1, column=1, value=f"As of {generated_at.strftime('%Y-%m-%d %H:%M')}")
    ws.cell(row=1, column=1).font = Font(italic=True)

    # Row 2: header
    for col_idx, header in enumerate(_HEADER_ROW, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9E1F2")
        cell.alignment = Alignment(horizontal="center")

    # Rows 3+: data (one row per config item, in config order — stable for external links)
    for row_idx, data_row in enumerate(rows, start=3):
        label, cny, usd, qty, price, as_of = data_row
        ws.cell(row=row_idx, column=_COL_LABEL + 1, value=label)
        ws.cell(row=row_idx, column=_COL_CNY + 1, value=cny)
        ws.cell(row=row_idx, column=_COL_USD + 1, value=usd)
        ws.cell(row=row_idx, column=_COL_QTY + 1, value=qty)
        ws.cell(row=row_idx, column=_COL_PRICE + 1, value=price)
        ws.cell(row=row_idx, column=_COL_ASOF + 1, value=as_of)

    # Column widths
    ws.column_dimensions["A"].width = 42
    for col in ["B", "C", "D", "E"]:
        ws.column_dimensions[col].width = 18
    ws.column_dimensions["F"].width = 14

    # ── 月度快照 sheet (WS-2) ────────────────────────────────────────────────
    ws_monthly = wb.create_sheet(title=MONTHLY_SHEET_NAME)
    for col_idx, header in enumerate(_MONTHLY_HEADER_ROW, start=1):
        cell = ws_monthly.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9E1F2")
        cell.alignment = Alignment(horizontal="center")

    for row_idx, data_row in enumerate(monthly_rows or [], start=2):
        for col_idx, value in enumerate(data_row, start=1):
            ws_monthly.cell(row=row_idx, column=col_idx, value=value)

    ws_monthly.column_dimensions["A"].width = 10
    ws_monthly.column_dimensions["B"].width = 22
    ws_monthly.column_dimensions["C"].width = 28
    ws_monthly.column_dimensions["D"].width = 16
    ws_monthly.column_dimensions["E"].width = 14
    ws_monthly.column_dimensions["F"].width = 18

    wb.save(str(out_path))


# ── Public entry point ────────────────────────────────────────────────────────

def export_reference_sheet(
    connector: Any,
    config: Dict[str, Any],
    out_dir: Optional[Path] = None,
    config_path: Optional[Path] = None,
) -> Path:
    """Generate UIS_Reference_Data.xlsx and write it to `out_dir`.

    Args:
        connector:   Active DatabaseConnector (read-only access — no writes).
        config:      Loaded settings dict (used to resolve default out_dir).
        out_dir:     Override output directory.  Defaults to config['finance_dir'].
        config_path: Override path to reference_sheet.yaml.  Defaults to
                     config/reference_sheet.yaml relative to cwd.

    Returns:
        Path to the written file.

    Raises:
        ValueError: if the safety guard trips (wrong filename or protected target).
        FileNotFoundError: if finance_dir is not configured and out_dir is not supplied.
    """
    # ── Resolve output directory ───────────────────────────────────────────
    if out_dir is None:
        finance_dir_str = config.get("finance_dir", "")
        if not finance_dir_str:
            raise FileNotFoundError(
                "export_reference_sheet: out_dir not supplied and "
                "finance_dir is not configured in settings"
            )
        out_dir = Path(finance_dir_str)

    out_dir = Path(out_dir)
    out_path = out_dir / OUTPUT_FILENAME

    # ── Safety guard (before ANY file I/O) ────────────────────────────────
    _assert_safe_output_path(out_path)

    # ── Load config ────────────────────────────────────────────────────────
    ref_config_path = config_path
    if ref_config_path is None:
        ref_config_path = _DEFAULT_CONFIG_PATH
    items = _load_reference_config(ref_config_path)

    # ── Query DB ───────────────────────────────────────────────────────────
    generated_at = datetime.now()
    rows = _build_rows(connector, items)

    # ── 月度快照 (WS-2): load prior history BEFORE we overwrite the file, compute
    #    fresh rows, then append-only-merge (see _merge_monthly_snapshot docstring). ──
    current_month = generated_at.strftime("%Y-%m")
    existing_monthly = _load_existing_monthly_snapshot(out_path)
    computed_monthly = _compute_monthly_snapshot_rows(
        connector, _MONTHLY_BACKFILL_START_MONTH, current_month,
    )
    monthly_rows = _merge_monthly_snapshot(existing_monthly, computed_monthly, current_month)

    # ── Ensure output directory exists ─────────────────────────────────────
    out_dir.mkdir(parents=True, exist_ok=True)

    # ── Write file ─────────────────────────────────────────────────────────
    _write_xlsx(out_path, rows, generated_at, monthly_rows)

    # Log a one-line summary
    non_blank = sum(1 for r in rows if any(v is not None for v in r[1:5]))
    logger.info(
        "Reference sheet written: %s (%d rows, %d non-blank)", out_path, len(rows), non_blank
    )
    return out_path
