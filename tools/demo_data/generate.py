#!/usr/bin/env python3
"""Deterministic synthetic demo-data generator for Huinsight (Program OSR, WS-1).

Reads tools/demo_data/persona.yaml and emits synthetic source files that are
byte-identical across runs given the same rng_seed. Every value is derived
from the persona spec + a seeded RNG — nothing here reads real owner data.

Phase 1: the 4 CSV fixtures —
    ibkr/IBKR_UIS_Report.csv
    ibkr_trades/IBKR_UIS_Report_trades.csv
    Individual-Positions-<snapshot>.csv
    Individual_XXX<tail>_Transactions_<snapshot>.csv
Phase 2 (this file, current scope): the 3 simple Excel workbooks —
    Gold_transactions.xlsx, Insurance_Portfolio.xlsx, RSU_transactions.xlsx
CN Fund and Financial Summary xlsx emitters are later phases.

Structure: one `emit_*` function per source file so later phases can add
emitters without touching this module's plumbing.

xlsx determinism: openpyxl stamps docProps/core.xml's `modified` field with
wall-clock time at save() (unconditionally — see openpyxl.writer.excel) and
zip members get wall-clock entry timestamps regardless of any properties set
beforehand. `_save_workbook()` patches both after save() so two runs with the
same rng_seed produce byte-identical .xlsx files.
"""
from __future__ import annotations

import argparse
import calendar
import random
import re
import shutil
import zipfile
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

import openpyxl
import yaml

HERE = Path(__file__).resolve().parent
PERSONA_PATH = HERE / "persona.yaml"
DEFAULT_OUT_DIR = HERE / "out"
FIXTURE_DIR = HERE.parent.parent / "tests" / "fixtures" / "readers"

# --install source -> glob(s) of generated file(s), relative to out_dir, to
# copy into tests/fixtures/readers/ at the SAME relative path (creating
# ibkr/ibkr_trades subdirs as needed). Schwab's positions/transactions
# filenames are date/account-tail derived (persona-controlled) rather than
# fixed, hence the glob instead of a literal name.
INSTALL_GLOBS: "dict[str, list[str]]" = {
    "ibkr": ["ibkr/IBKR_UIS_Report.csv", "ibkr_trades/IBKR_UIS_Report_trades.csv"],
    "schwab": ["Individual-Positions-*.csv", "Individual_*_Transactions_*.csv"],
    "gold": ["Gold_transactions.xlsx"],
    "insurance": ["Insurance_Portfolio.xlsx"],
    "rsu": ["RSU_transactions.xlsx"],
    "cn_fund": ["funding_transactions.xlsx"],
    "financial_summary": ["Financial_Summary_new.xlsx"],
}


# ---------------------------------------------------------------------------
# Persona loading
# ---------------------------------------------------------------------------


def load_persona(path: Path = PERSONA_PATH) -> dict[str, Any]:
    with open(path, "r", encoding="utf-8") as fh:
        return yaml.safe_load(fh)


# ---------------------------------------------------------------------------
# Formatting helpers (mirror the exact string shapes the real exports use —
# see docs/plans/2026-08-16-ws0-shape-manifest.md).
# ---------------------------------------------------------------------------


def _fmt_num(x: float) -> str:
    """Bare number: integers unadorned, else 2 decimal places (IBKR flex style)."""
    if abs(x - round(x)) < 1e-9:
        return str(int(round(x)))
    return f"{x:.2f}"


def _fmt_qty(x: float) -> str:
    """Schwab Quantity column: bare number, up to 4dp, trailing zeros trimmed."""
    if abs(x - round(x)) < 1e-9:
        return str(int(round(x)))
    s = f"{x:.4f}".rstrip("0").rstrip(".")
    return s


def _fmt_dollar(x: float) -> str:
    sign = "-" if x < -1e-9 else ""
    return f"${abs(x):,.2f}" if not sign else f"-${abs(x):,.2f}"


def _fmt_percent(x: float) -> str:
    sign = "-" if x < -1e-9 else ""
    return f"{sign}{abs(x):.2f}%"


def _add_months(d: date, months: int) -> date:
    total = d.year * 12 + (d.month - 1) + months
    year, month = divmod(total, 12)
    month += 1
    day = min(d.day, calendar.monthrange(year, month)[1])
    return date(year, month, day)


def _spread_dates(rng: random.Random, start: date, end: date, n: int) -> list[date]:
    """n dates roughly evenly spaced between start and end, with jitter."""
    span_days = (end - start).days
    if n <= 0:
        return []
    if n == 1:
        return [start + timedelta(days=span_days // 2)]
    step = span_days / (n - 1)
    out = []
    for i in range(n):
        base = start + timedelta(days=round(step * i))
        jitter = timedelta(days=rng.randint(-3, 3))
        d = base + jitter
        if d < start:
            d = start
        if d > end:
            d = end
        out.append(d)
    return out


def _csv_row_all_quoted(fields: list[str]) -> str:
    """Comma-join fields, each wrapped in double quotes (IBKR flex style)."""
    return ",".join(f'"{f}"' for f in fields)


def _write_lines(path: Path, lines: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8", newline="") as fh:
        fh.write("\n".join(lines) + "\n")


# ---------------------------------------------------------------------------
# xlsx determinism helpers
# ---------------------------------------------------------------------------

_XLSX_FIXED_ZIP_DATE = (2026, 1, 1, 0, 0, 0)


def _fixed_xlsx_datetime(persona: dict[str, Any]) -> datetime:
    d = persona["timeline"]["snapshot_date"]
    return datetime(d.year, d.month, d.day)


def _dt(d: date) -> datetime:
    """date -> datetime at midnight, for writing into Excel date cells."""
    return datetime(d.year, d.month, d.day)


def _freeze_xlsx(path: Path, fixed_dt: datetime) -> None:
    """Rewrite an .xlsx (a zip) so two runs produce byte-identical files.

    openpyxl.writer.excel unconditionally sets docProps/core.xml's
    `modified` field to datetime.now() at save() time, and zip member
    entries get wall-clock timestamps regardless of workbook.properties.
    Patch the modified field and freeze every zip entry's date_time.
    """
    fixed_iso = fixed_dt.strftime("%Y-%m-%dT%H:%M:%SZ").encode()
    tmp = path.with_suffix(path.suffix + ".tmp")
    with zipfile.ZipFile(path, "r") as zin, zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename == "docProps/core.xml":
                data = re.sub(
                    rb"(<dcterms:modified[^>]*>)[^<]*(</dcterms:modified>)",
                    lambda m: m.group(1) + fixed_iso + m.group(2),
                    data,
                )
            item.date_time = _XLSX_FIXED_ZIP_DATE
            zout.writestr(item, data)
    shutil.move(str(tmp), str(path))


def _save_workbook(wb: "openpyxl.Workbook", path: Path, fixed_dt: datetime) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.properties.created = fixed_dt
    wb.properties.modified = fixed_dt
    wb.save(path)
    _freeze_xlsx(path, fixed_dt)


# ---------------------------------------------------------------------------
# IBKR Flex CSVs
# ---------------------------------------------------------------------------


def _ibkr_symbol_lookup(persona: dict[str, Any]) -> dict[str, str]:
    return {p["symbol"]: p["description"] for p in persona["ibkr"]["positions"]}


def _emit_ibkr_report(
    *,
    account: str,
    name_en: str,
    report_date: date,
    positions: list[dict[str, Any]],
    trades: list[dict[str, Any]],
    transfers: list[tuple[str, float]],
    symbol_desc: dict[str, str],
    cash_usd: float,
    rng: random.Random,
) -> list[str]:
    """Build one Flex CSV's lines (list of section rows, all-quoted)."""
    lines: list[str] = []

    period_from = date(report_date.year, 1, 1)
    generated = report_date + timedelta(days=1)
    generated_time = f"{rng.randint(6, 21):02d}{rng.randint(0, 59):02d}{rng.randint(0, 59):02d}"

    lines.append(_csv_row_all_quoted(
        ["BOF", account, "IBKR_UIS_Report", "5",
         period_from.isoformat(), report_date.isoformat(),
         f"{generated.isoformat()};{generated_time}", "100", "100"]
    ))
    lines.append(_csv_row_all_quoted(["BOA", account]))

    # --- ACCT ---
    lines.append(_csv_row_all_quoted(["BOS", "ACCT", "Account Information"]))
    lines.append(_csv_row_all_quoted(
        ["ClientAccountID", "AccountAlias", "CurrencyPrimary", "Name"]
    ))
    lines.append(_csv_row_all_quoted([account, "", "USD", name_en]))
    lines.append(_csv_row_all_quoted(["EOS", "ACCT", "1", "0"]))

    # --- CRTT (cash report) ---
    lines.append(_csv_row_all_quoted(["BOS", "CRTT", "Cash Report; trade date basis"]))
    lines.append(_csv_row_all_quoted([
        "ClientAccountID", "CurrencyPrimary", "AccountTransfers",
        "AccountTransfersSecurities", "AccountTransfersCommodities",
        "EndingCash", "EndingCashSecurities", "EndingCashCommodities",
    ]))
    lines.append(_csv_row_all_quoted([
        account, "BASE_SUMMARY", "0", "0", "0",
        _fmt_num(cash_usd), _fmt_num(cash_usd), "0",
    ]))
    lines.append(_csv_row_all_quoted(["EOS", "CRTT", "1", _fmt_num(cash_usd)]))

    # --- POST (positions) ---
    lines.append(_csv_row_all_quoted(["BOS", "POST", "Position; trade date basis"]))
    lines.append(_csv_row_all_quoted([
        "ClientAccountID", "AccountAlias", "CurrencyPrimary", "Symbol", "Description",
        "ReportDate", "Quantity", "MarkPrice", "PositionValue", "PositionValueInBase",
        "CostBasisPrice", "CostBasisMoney", "FifoPnlUnrealized",
        "UnrealizedCapitalGainsPnl", "UnrealizedFxPnl", "HoldingPeriodDateTime",
    ]))
    post_total_value = 0.0
    for p in positions:
        qty = p["qty"]
        mark_price = p["mark_price"]
        cost_price = p["cost_price"]
        position_value = qty * mark_price
        cost_basis_money = qty * cost_price
        fifo_pnl = position_value - cost_basis_money
        post_total_value += position_value
        lines.append(_csv_row_all_quoted([
            account, "", "USD", p["symbol"], symbol_desc[p["symbol"]],
            report_date.isoformat(), _fmt_num(qty), _fmt_num(mark_price),
            _fmt_num(position_value), _fmt_num(position_value),
            _fmt_num(cost_price), _fmt_num(cost_basis_money), _fmt_num(fifo_pnl),
            "0", "0", "",
        ]))
    lines.append(_csv_row_all_quoted(["EOS", "POST", str(len(positions)), _fmt_num(post_total_value)]))

    # --- TRNT (trades) ---
    lines.append(_csv_row_all_quoted(["BOS", "TRNT", "Trades; trade date basis"]))
    lines.append(_csv_row_all_quoted([
        "ClientAccountID", "AccountAlias", "CurrencyPrimary", "Symbol", "Description",
        "DateTime", "TradeDate", "Quantity", "TradePrice", "IBCommission",
        "IBCommissionCurrency", "Buy/Sell", "OrderType",
    ]))
    if trades:
        for t in trades:
            lines.append(_csv_row_all_quoted([
                account, "", "USD", t["symbol"], symbol_desc[t["symbol"]],
                f'{report_date.isoformat()};{t["time_suffix"]}', report_date.isoformat(),
                _fmt_num(t["quantity"]), _fmt_num(t["trade_price"]),
                _fmt_num(t["commission"]), "USD", t["side"], "LMT",
            ]))
        lines.append(_csv_row_all_quoted(["EOS", "TRNT", str(len(trades)), "0"]))
    else:
        lines.append(_csv_row_all_quoted(["EOS", "TRNT", "0", "0"]))

    # --- TRFR (transfers) ---
    lines.append(_csv_row_all_quoted(["BOS", "TRFR", "Transfers"]))
    lines.append(_csv_row_all_quoted([
        "ClientAccountID", "AccountAlias", "CurrencyPrimary", "Symbol", "ReportDate",
        "DateTime", "SettleDate", "Type", "Direction", "TransferCompany",
        "TransferAccount", "TransferAccountName", "Quantity", "TransferPrice",
        "PositionAmount", "PositionAmountInBase", "CapitalGainsPnl", "FxPnl",
        "PnlAmount", "PnlAmountInBase", "CashTransfer",
    ]))
    if transfers:
        transfer_date = report_date - timedelta(days=4)
        settle_date = transfer_date + timedelta(days=1)
        cost_price_by_symbol = {p["symbol"]: p["cost_price"] for p in positions}
        total_amount = 0.0
        for symbol, qty in transfers:
            cost_price = cost_price_by_symbol[symbol]
            amount = qty * cost_price
            total_amount += amount
            lines.append(_csv_row_all_quoted([
                account, "", "USD", symbol,
                transfer_date.isoformat(), transfer_date.isoformat(), settle_date.isoformat(),
                "ACATS", "IN", "--", "88881234", "",
                _fmt_num(qty), "0", _fmt_num(amount), _fmt_num(amount),
                "0", "0", "0", "0", "0",
            ]))
        lines.append(_csv_row_all_quoted(["EOS", "TRFR", str(len(transfers)), _fmt_num(total_amount)]))
    else:
        lines.append(_csv_row_all_quoted(["EOS", "TRFR", "0", "0"]))

    lines.append(_csv_row_all_quoted(["EOA", account]))
    lines.append(_csv_row_all_quoted(["EOF"]))
    return lines


def emit_ibkr(persona: dict[str, Any], out_dir: Path, rng: random.Random) -> list[Path]:
    account = persona["identity"]["ibkr_account"]
    name_en = persona["identity"]["name_en"]
    symbol_desc = _ibkr_symbol_lookup(persona)
    cash_usd = persona["ibkr"]["cash_usd"]

    report_dates = persona["timeline"]["ibkr_report_dates"]
    report1_date, report2_date = report_dates[0], report_dates[1]

    # Report 1: positions established via transfer, no trades yet.
    positions1 = []
    for p in persona["ibkr"]["positions"]:
        cost_price = p["cost_usd"] / p["qty"]
        mark_price = round(cost_price * rng.uniform(0.92, 1.12), 2)
        positions1.append({
            "symbol": p["symbol"], "qty": p["qty"],
            "cost_price": round(cost_price, 4), "mark_price": mark_price,
        })

    n_transfer_count = persona["ibkr"]["transfers"]["count"]
    # Split the first position's quantity across enough rows to hit the
    # persona-declared transfer count exactly; remaining positions get one
    # row each.
    positions_syms = [p["symbol"] for p in persona["ibkr"]["positions"]]
    transfers: list[tuple[str, float]] = []
    extra_rows_needed = max(0, n_transfer_count - len(positions_syms))
    first_symbol = positions_syms[0]
    first_qty = next(p["qty"] for p in positions1 if p["symbol"] == first_symbol)
    if extra_rows_needed:
        # Split first symbol's quantity into (extra_rows_needed + 1) parts.
        parts = extra_rows_needed + 1
        base_part = first_qty // parts
        remainder = first_qty - base_part * (parts - 1)
        for i in range(parts - 1):
            transfers.append((first_symbol, base_part))
        transfers.append((first_symbol, remainder))
        for sym in positions_syms[1:]:
            qty = next(p["qty"] for p in positions1 if p["symbol"] == sym)
            transfers.append((sym, qty))
    else:
        for sym in positions_syms:
            qty = next(p["qty"] for p in positions1 if p["symbol"] == sym)
            transfers.append((sym, qty))

    report1_lines = _emit_ibkr_report(
        account=account, name_en=name_en, report_date=report1_date,
        positions=positions1, trades=[], transfers=transfers,
        symbol_desc=symbol_desc, cash_usd=cash_usd, rng=rng,
    )

    # Report 2: apply persona.ibkr.trades on top of report-1 positions.
    positions1_by_symbol = {p["symbol"]: p for p in positions1}
    trades_cfg = persona["ibkr"]["trades"]
    prepared_trades = []
    for t in trades_cfg:
        side = t["side"].upper()
        qty = t["qty"]
        prior = positions1_by_symbol[t["symbol"]]
        trade_price = round(prior["cost_price"] * rng.uniform(0.95, 1.2), 2)
        commission = round(rng.uniform(-1.0, -0.1), 2)
        time_suffix = f"{rng.randint(9, 15):02d}{rng.randint(0, 59):02d}{rng.randint(0, 59):02d}"
        prepared_trades.append({
            "symbol": t["symbol"],
            "quantity": qty,  # signed: SELL rows carry a negative quantity
            "trade_price": trade_price,
            "commission": commission,
            "side": side,
            "time_suffix": time_suffix,
        })

    positions2 = []
    for sym, prior in positions1_by_symbol.items():
        net_qty_change = sum(
            t["quantity"] for t in prepared_trades if t["symbol"] == sym
        )
        new_qty = prior["qty"] + net_qty_change
        buys = [t for t in prepared_trades if t["symbol"] == sym and t["quantity"] > 0]
        if buys and new_qty > 0:
            bought_cost = sum(t["quantity"] * t["trade_price"] for t in buys)
            new_cost_price = (
                (prior["qty"] * prior["cost_price"] + bought_cost) / new_qty
            )
        else:
            new_cost_price = prior["cost_price"]
        mark_price = round(prior["mark_price"] * rng.uniform(0.98, 1.05), 2)
        positions2.append({
            "symbol": sym, "qty": new_qty,
            "cost_price": round(new_cost_price, 4), "mark_price": mark_price,
        })

    # A small decorative TRFR row, matching the real report2's own shape (its
    # TRFR carried one leftover transfer event not reconciled against POST —
    # a real Flex-report quirk, not a bug to "fix"). Exercises TRFR alongside
    # TRNT in the same report so a trades-bearing fixture also covers transfers.
    report2_transfers = [(positions_syms[0], 5.0)]

    report2_lines = _emit_ibkr_report(
        account=account, name_en=name_en, report_date=report2_date,
        positions=positions2, trades=prepared_trades, transfers=report2_transfers,
        symbol_desc=symbol_desc, cash_usd=cash_usd, rng=rng,
    )

    p1 = out_dir / "ibkr" / "IBKR_UIS_Report.csv"
    p2 = out_dir / "ibkr_trades" / "IBKR_UIS_Report_trades.csv"
    _write_lines(p1, report1_lines)
    _write_lines(p2, report2_lines)
    return [p1, p2]


# ---------------------------------------------------------------------------
# Schwab positions CSV
# ---------------------------------------------------------------------------

_SCHWAB_POSITIONS_HEADER = [
    "Symbol", "Description", "Qty (Quantity)", "Price",
    "Price Chng $ (Price Change $)", "Price Chng % (Price Change %)",
    "Mkt Val (Market Value)", "Day Chng $ (Day Change $)", "Day Chng % (Day Change %)",
    "Cost Basis", "Gain $ (Gain/Loss $)", "Gain % (Gain/Loss %)",
    "Reinvest?", "Reinvest Capital Gains?", "Asset Type",
]


def _schwab_row(fields: list[str]) -> str:
    """Positions CSV row style: every field quoted, PLUS a trailing bare comma."""
    return ",".join(f'"{f}"' for f in fields) + ","


def emit_schwab_positions(persona: dict[str, Any], out_dir: Path, rng: random.Random) -> Path:
    tail = persona["identity"]["schwab_account_tail"]
    snapshot_date: date = persona["timeline"]["snapshot_date"]
    positions = persona["schwab"]["positions"]
    cash_usd = persona["schwab"]["cash_usd"]

    lines: list[str] = []
    preamble = (
        f"Positions for account Individual ...{tail} as of "
        f"06:04 AM ET, {snapshot_date.year:04d}/{snapshot_date.month:02d}/{snapshot_date.day:02d}"
    )
    lines.append(f'"{preamble}"')
    lines.append("")
    lines.append(_schwab_row(_SCHWAB_POSITIONS_HEADER))

    total_mkt_val = 0.0
    total_cost_basis = 0.0
    total_gain = 0.0
    for p in positions:
        qty = p["qty"]
        cost_price = p["cost_usd"] / qty
        price = round(cost_price * rng.uniform(0.9, 1.35), 2)
        mkt_val = round(qty * price, 2)
        price_chng = round(rng.uniform(-5, 5), 2)
        price_chng_pct = round((price_chng / price) * 100, 2) if price else 0.0
        day_chng = round(rng.uniform(-mkt_val * 0.02, mkt_val * 0.02), 2)
        day_chng_pct = round((day_chng / mkt_val) * 100, 2) if mkt_val else 0.0
        cost_basis = round(p["cost_usd"], 2)
        gain = round(mkt_val - cost_basis, 2)
        gain_pct = round((gain / cost_basis) * 100, 2) if cost_basis else 0.0

        total_mkt_val += mkt_val
        total_cost_basis += cost_basis
        total_gain += gain

        lines.append(_schwab_row([
            p["symbol"], p["description"], _fmt_qty(qty), f"{price:.2f}",
            _fmt_dollar(price_chng), _fmt_percent(price_chng_pct),
            _fmt_dollar(mkt_val), _fmt_dollar(day_chng), _fmt_percent(day_chng_pct),
            _fmt_dollar(cost_basis), _fmt_dollar(gain), _fmt_percent(gain_pct),
            "No", "N/A", p["type"],
        ]))

    lines.append(_schwab_row([
        "Cash & Cash Investments", "--", "--", "--", "--", "--",
        _fmt_dollar(cash_usd), "$0.00", "0%", "--", "--", "--", "--", "--",
        "Cash and Money Market",
    ]))

    total_mkt_val += cash_usd
    lines.append(_schwab_row([
        "Positions Total", "", "--", "--", "--", "--",
        _fmt_dollar(total_mkt_val), _fmt_dollar(0.0), _fmt_percent(0.0),
        _fmt_dollar(total_cost_basis), _fmt_dollar(total_gain),
        _fmt_percent((total_gain / total_cost_basis) * 100 if total_cost_basis else 0.0),
        "--", "--", "--",
    ]))

    out_path = out_dir / f"Individual-Positions-{snapshot_date.strftime('%Y-%m-%d')}-060406.csv"
    _write_lines(out_path, lines)
    return out_path


# ---------------------------------------------------------------------------
# Schwab transactions CSV
# ---------------------------------------------------------------------------

_SCHWAB_TXN_HEADER = ["Date", "Action", "Symbol", "Description", "Quantity", "Price", "Fees & Comm", "Amount"]


def _schwab_txn_row(fields: list[str]) -> str:
    return ",".join(f'"{f}"' for f in fields)


def _schwab_txn_symbol(symbol: str) -> str:
    """Transactions CSV historically drops the slash Schwab shows in Positions
    (BRK/B -> BRKB); the bare compound ticker exercises the symbol_norm path."""
    return symbol.replace("/", "")


def emit_schwab_transactions(persona: dict[str, Any], out_dir: Path, rng: random.Random) -> tuple[Path, list]:
    tail = persona["identity"]["schwab_account_tail"]
    snapshot_date: date = persona["timeline"]["snapshot_date"]
    positions = persona["schwab"]["positions"]
    txn_cfg = persona["schwab"]["transactions"]
    months = txn_cfg["months_of_history"]
    start_date = _add_months(snapshot_date, -months)

    rows: list[tuple[date, str, str, str, str, str, str, str]] = []

    def add(d: date, action: str, symbol: str, description: str,
            qty: str = "", price: str = "", fees: str = "", amount: str = "") -> None:
        rows.append((d, action, symbol, description, qty, price, fees, amount))

    # --- Buy: initial buy establishing each position ---
    buy_dates = _spread_dates(rng, start_date, snapshot_date - timedelta(days=30), len(positions))
    cost_price_by_symbol: dict[str, float] = {}
    for pos, d in zip(positions, buy_dates):
        cost_price = round(pos["cost_usd"] / pos["qty"], 4)
        cost_price_by_symbol[pos["symbol"]] = cost_price
        amount = -(cost_price * pos["qty"])
        add(
            d, "Buy", _schwab_txn_symbol(pos["symbol"]), pos["description"],
            _fmt_qty(pos["qty"]), f"{cost_price:.4f}", "", _fmt_dollar(amount),
        )

    # --- Sell: a couple of partial sells for vocab coverage ---
    sell_targets = positions[:2]
    sell_dates = _spread_dates(rng, start_date, snapshot_date - timedelta(days=10), len(sell_targets))
    for pos, d in zip(sell_targets, sell_dates):
        cost_price = cost_price_by_symbol[pos["symbol"]]
        sell_qty = round(pos["qty"] * rng.uniform(0.05, 0.15), 4)
        sell_price = round(cost_price * rng.uniform(1.0, 1.2), 2)
        amount = sell_qty * sell_price
        fees = round(rng.uniform(0.01, 0.1), 2)
        add(
            d, "Sell", _schwab_txn_symbol(pos["symbol"]), pos["description"],
            _fmt_qty(sell_qty), f"{sell_price:.2f}", _fmt_dollar(fees), _fmt_dollar(amount),
        )

    # --- Dividend cycles: NRA Tax Adj + Cash Dividend pairs, 3 cycles/symbol ---
    for pos in positions:
        cycle_dates = _spread_dates(rng, start_date, snapshot_date - timedelta(days=5), 3)
        for d in cycle_dates:
            div_amount = round(rng.uniform(5, 250), 2)
            tax_amount = round(div_amount * rng.uniform(0.08, 0.15), 2)
            sym = _schwab_txn_symbol(pos["symbol"])
            add(d, "NRA Tax Adj", sym, pos["description"], amount=_fmt_dollar(-tax_amount))
            add(d, "Cash Dividend", sym, pos["description"], amount=_fmt_dollar(div_amount))

    # --- Remaining required_actions vocabulary (each >= 1 row) ---
    required = txn_cfg["required_actions"]
    vocab_symbol = positions[2]["symbol"]  # SCHD — arbitrary stable choice
    vocab_desc = positions[2]["description"]
    vocab_cost_price = cost_price_by_symbol[vocab_symbol]

    if "Qualified Dividend" in required:
        d = _spread_dates(rng, start_date, snapshot_date, 1)[0]
        add(d, "Qualified Dividend", _schwab_txn_symbol(positions[3]["symbol"]),
            positions[3]["description"], amount=_fmt_dollar(round(rng.uniform(5, 80), 2)))

    if "Credit Interest" in required:
        d = _spread_dates(rng, start_date, snapshot_date, 1)[0]
        add(d, "Credit Interest", "", "Bank Sweep Interest",
            amount=_fmt_dollar(round(rng.uniform(0.01, 5), 2)))

    wire_transfers: list[dict[str, Any]] = []
    if "Wire Transfer" in required:
        d = _spread_dates(rng, start_date, snapshot_date, 1)[0]
        wire_amount = round(rng.uniform(2000, 20000), 2)
        add(d, "Wire Transfer", "", f"Wire Funds Received, {persona['identity']['name_en'].upper()}",
            amount=_fmt_dollar(wire_amount))
        wire_transfers.append({"date": d, "amount_usd": wire_amount})

    if "Journal" in required:
        d = _spread_dates(rng, start_date, snapshot_date, 1)[0]
        add(d, "Journal", "", "Journaled Cash", amount=_fmt_dollar(round(rng.uniform(50, 500), 2)))

    if "Reinvest Dividend" in required:
        d = _spread_dates(rng, start_date, snapshot_date, 1)[0]
        reinvest_amount = round(rng.uniform(10, 60), 2)
        reinvest_price = round(vocab_cost_price * rng.uniform(1.0, 1.15), 2)
        reinvest_qty = round(reinvest_amount / reinvest_price, 4)
        add(d, "Reinvest Dividend", _schwab_txn_symbol(vocab_symbol), vocab_desc,
            qty=_fmt_qty(reinvest_qty), price=f"{reinvest_price:.2f}", amount=_fmt_dollar(reinvest_amount))

    if "Reinvest Shares" in required:
        d = _spread_dates(rng, start_date, snapshot_date, 1)[0]
        reinvest_amount = round(rng.uniform(10, 60), 2)
        reinvest_price = round(vocab_cost_price * rng.uniform(1.0, 1.15), 2)
        reinvest_qty = round(reinvest_amount / reinvest_price, 4)
        add(d, "Reinvest Shares", _schwab_txn_symbol(vocab_symbol), vocab_desc,
            qty=_fmt_qty(reinvest_qty), price=f"{reinvest_price:.2f}", amount=_fmt_dollar(reinvest_amount))

    if "Security Transfer" in required:
        d_in, d_out = _spread_dates(rng, start_date, snapshot_date, 2)
        transfer_symbol = positions[4]["symbol"]
        transfer_desc = positions[4]["description"]
        add(d_in, "Security Transfer", _schwab_txn_symbol(transfer_symbol), transfer_desc,
            qty=_fmt_qty(round(rng.uniform(5, 20), 4)))
        add(d_out, "Security Transfer", _schwab_txn_symbol(transfer_symbol), transfer_desc,
            qty=_fmt_qty(-round(rng.uniform(5, 20), 4)))

    for action in required:
        assert action in {r[1] for r in rows}, f"required_action never emitted: {action}"
    assert any(
        r[2] == "BRKB" for r in rows
    ), "required_symbols BRKB never emitted"

    rows.sort(key=lambda r: r[0], reverse=True)

    lines = [_schwab_txn_row(_SCHWAB_TXN_HEADER)]
    for d, action, symbol, description, qty, price, fees, amount in rows:
        lines.append(_schwab_txn_row([
            d.strftime("%m/%d/%Y"), action, symbol, description, qty, price, fees, amount,
        ]))

    out_path = out_dir / f"Individual_XXX{tail}_Transactions_{snapshot_date.strftime('%Y%m%d')}-060417.csv"
    _write_lines(out_path, lines)
    return out_path, wire_transfers


# ---------------------------------------------------------------------------
# Gold (Excel) — Gold_transactions.xlsx
# ---------------------------------------------------------------------------

_GOLD_HOLDINGS_HEADER = ["资产类别", "标的名称", "持有数量", "单位", "平均成本价", "单价", "当前市值", "未实现盈亏", "交易账户"]
_GOLD_TXN_HEADER = ["交易日期", "资产类别", "标的名称", "交易类型", "金额", "数量", "价格", "手续费", "交易账户"]


def emit_gold(persona: dict[str, Any], out_dir: Path, rng: random.Random) -> tuple[Path, list]:
    accounts = persona["gold"]["holdings"]
    ref_price = persona["timeline"]["gold_cny_per_gram"]
    snapshot_date: date = persona["timeline"]["snapshot_date"]
    required_types = persona["gold"]["transactions"]["required_types"]
    approx_rows = persona["gold"]["transactions"]["approx_rows"]
    start_date = _add_months(snapshot_date, -18)

    n_accounts = len(accounts)
    rows_per_account = max(4, approx_rows // n_accounts)

    all_txns: list[dict[str, Any]] = []
    holdings_rows: list[dict[str, Any]] = []

    for acct in accounts:
        target_grams = acct["grams"]
        account_name = acct["account"]
        asset_name = acct["asset"]

        n_buys = max(2, rows_per_account - 2)
        dates = _spread_dates(rng, start_date, snapshot_date - timedelta(days=5), n_buys + 2)

        price = ref_price
        running_qty = 0.0
        buy_cost_total = 0.0
        buy_qty_total = 0.0
        acct_txns: list[dict[str, Any]] = []

        for d in dates[:n_buys]:
            price = round(price * rng.uniform(0.97, 1.04), 2)
            qty = round(rng.uniform(3, 15), 4)
            amount = round(qty * price, 2)
            fees = round(rng.uniform(0, 0.15), 4) if rng.random() > 0.4 else 0.0
            acct_txns.append({"date": d, "type": "买入", "amount": amount, "qty": qty, "price": price, "fees": fees})
            running_qty += qty
            buy_cost_total += amount
            buy_qty_total += qty

        d_sell = dates[n_buys]
        price = round(price * rng.uniform(0.97, 1.04), 2)
        sell_qty = round(max(min(running_qty * rng.uniform(0.1, 0.3), running_qty - 1), 0.5), 4)
        amount = round(sell_qty * price, 2)
        acct_txns.append({"date": d_sell, "type": "卖出", "amount": amount, "qty": sell_qty, "price": price, "fees": 0.0})
        running_qty -= sell_qty

        d_int = dates[n_buys + 1]
        price = round(price * rng.uniform(0.97, 1.04), 2)
        int_qty = round(rng.uniform(0.0005, 0.01), 4)
        amount = round(int_qty * price, 2)
        acct_txns.append({"date": d_int, "type": "结息", "amount": amount, "qty": int_qty, "price": price, "fees": 0.0})
        running_qty += int_qty

        residual = round(target_grams - running_qty, 4)
        if abs(residual) > 1e-6:
            price = round(price * rng.uniform(0.98, 1.02), 2)
            plug_type = "买入" if residual > 0 else "卖出"
            plug_qty = abs(residual)
            amount = round(plug_qty * price, 2)
            acct_txns.append({
                "date": snapshot_date - timedelta(days=2), "type": plug_type,
                "amount": amount, "qty": plug_qty, "price": price, "fees": 0.0,
            })
            if plug_type == "买入":
                buy_cost_total += amount
                buy_qty_total += plug_qty
            running_qty += residual

        for t in acct_txns:
            t["account"] = account_name
            t["asset"] = asset_name
            all_txns.append(t)

        avg_cost = round(buy_cost_total / buy_qty_total, 2) if buy_qty_total else price
        holdings_rows.append({
            "asset_name": asset_name, "qty": round(running_qty, 4),
            "avg_cost": avg_cost, "unit_price": price,
            "market_value": round(running_qty * price, 2),
            "unrealized_pnl": round(running_qty * price - running_qty * avg_cost, 2),
            "account": account_name,
        })

    for rt in required_types:
        assert any(t["type"] == rt for t in all_txns), f"gold required_type never emitted: {rt}"

    all_txns.sort(key=lambda t: t["date"])

    wb = openpyxl.Workbook()
    ws_h = wb.active
    ws_h.title = "黄金持仓"
    ws_h.append(_GOLD_HOLDINGS_HEADER)
    for h in holdings_rows:
        ws_h.append([
            "黄金", h["asset_name"], h["qty"], "克", h["avg_cost"], h["unit_price"],
            h["market_value"], h["unrealized_pnl"], h["account"],
        ])

    ws_t = wb.create_sheet("黄金交易记录")
    ws_t.append(_GOLD_TXN_HEADER)
    for t in all_txns:
        ws_t.append([_dt(t["date"]), "黄金", t["asset"], t["type"], t["amount"], t["qty"], t["price"], t["fees"], t["account"]])

    out_path = out_dir / "Gold_transactions.xlsx"
    _save_workbook(wb, out_path, _fixed_xlsx_datetime(persona))
    return out_path, all_txns


# ---------------------------------------------------------------------------
# Insurance (Excel) — Insurance_Portfolio.xlsx
# ---------------------------------------------------------------------------

_INSURANCE_SUMMARY_HEADER = [
    "产品名称", "保险公司", "产品类型", "开始日期", "保障期限", "缴费期限", "年保费", "保额",
    "保障范围", "保单状态", "现金价值", "被保险人", "投保渠道", "备注",
]
_INSURANCE_FAMILY_HEADER = [
    "家庭成员", "险种分类", "保险公司", "产品/福利名称", "基本保额", "保单开始时间(生效日)",
    "保障期间", "缴费频次", "保费金额(元)", "社保/医保基础", "备注/核心权益",
]
_INSURANCE_LAPSED_HEADER = [
    "产品名称", "保险公司", "产品类型", "被保险人", "投保日期", "退保日期", "持有年数",
    "已缴保费总额", "退保拿到现金价值", "净损失", "退保原因", "经验教训",
]


def emit_insurance(persona: dict[str, Any], out_dir: Path, rng: random.Random) -> tuple[Path, dict]:
    policies = persona["insurance"]["policies"]
    premium_cols = persona["insurance"]["premium_record_columns"]
    insured = persona["insurance"]["insured"]
    snapshot_date: date = persona["timeline"]["snapshot_date"]
    start_month: str = persona["timeline"]["start_month"]
    year, month = (int(x) for x in start_month.split("-"))
    first_month = date(year, month, 1)

    wb = openpyxl.Workbook()

    # --- 保险汇总 (holdings) ---
    ws_s = wb.active
    ws_s.title = "保险汇总"
    ws_s.append(_INSURANCE_SUMMARY_HEADER)

    coverage_scope_by_type = {
        "综合(寿险+重疾)": "终身寿+重疾保障",
        "定期重疾": "定期重大疾病保障",
        "意外险": "意外身故/伤残保障",
    }
    channel_by_index = ["线下保险代理", "互联网直销", "线下保险代理"]
    for i, p in enumerate(policies):
        is_active = p["status"] != "退保"
        cash_value = round(rng.uniform(8000, 25000), 2) if (is_active and i == 0) else 0.0
        coverage_term = "终身" if i == 0 else ("30年" if is_active else "1年")
        payment_term = _dt(date(p["start"].year + 20, p["start"].month, p["start"].day)) if is_active else "已退保"
        scope = coverage_scope_by_type.get(p["type"], f"{p['type']}保障")
        ws_s.append([
            p["product"], p["insurer"], p["type"], _dt(p["start"]), coverage_term, payment_term,
            p["annual_premium_cny"], p["coverage_cny"], scope, p["status"], cash_value,
            insured, channel_by_index[i % len(channel_by_index)],
            f"示例保单，用于演示数据生成 #{i + 1}",
        ])

    assert any(p["status"] == "退保" for p in policies), "no 退保 policy present in persona"

    # --- 保费记录 (premiums, wide melt) ---
    ws_p = wb.create_sheet("保费记录")
    ws_p.append(["日期  "] + list(premium_cols))
    # Same month count as the FS balance_sheet sheet (persona's overall window).
    months = persona["financial_summary"]["balance_sheet"]["months"]
    lump_sum_month = rng.randint(0, min(11, months - 1))
    premium_by_month: dict[date, dict[str, float]] = {}
    for i in range(months):
        d = _add_months(first_month, i)
        row = [_dt(d)]
        month_amounts: dict[str, float] = {}
        for col_idx, col in enumerate(premium_cols):
            if col_idx == 0:
                # annual whole-life premium: one lump sum per year, else 0
                policy = policies[0]
                val = policy["annual_premium_cny"] if (i % 12) == lump_sum_month else 0
            elif col_idx == 1:
                # employer group insurance: no employee-paid premium
                val = 0
            else:
                # small recurring monthly premium, with a couple of blanks
                val = None if rng.random() < 0.08 else round(rng.uniform(80, 220), 2)
            row.append(val)
            month_amounts[col] = val or 0.0
        premium_by_month[d] = month_amounts
        ws_p.append(row)

    # --- 全家保险 (not consumed by the reader; shape realism only) ---
    ws_f = wb.create_sheet("全家保险")
    ws_f.append(_INSURANCE_FAMILY_HEADER)
    ws_f.append([
        insured, "寿险底仓", policies[0]["insurer"], policies[0]["product"], f"{policies[0]['coverage_cny']}元",
        _dt(policies[0]["start"]), "终身", "年缴", policies[0]["annual_premium_cny"], "有社保", "示例家庭保单一览行",
    ])
    ws_f.append([
        insured, "定期重疾", policies[1]["insurer"], policies[1]["product"], f"{policies[1]['coverage_cny']}元",
        _dt(policies[1]["start"]), "定期", "年缴", policies[1]["annual_premium_cny"], "有社保", "示例家庭保单二览行",
    ])

    # --- 退保历史 (not consumed by the reader; shape realism only) ---
    ws_l = wb.create_sheet("退保历史")
    ws_l.append(_INSURANCE_LAPSED_HEADER)
    lapsed = policies[-1]
    held_years = max(1, snapshot_date.year - lapsed["start"].year)
    paid_total = round(lapsed["annual_premium_cny"] * held_years, 2)
    cash_back = round(paid_total * 0.05, 2)
    ws_l.append([
        lapsed["product"], lapsed["insurer"], lapsed["type"], insured, _dt(lapsed["start"]),
        _dt(snapshot_date), f"{held_years}年", paid_total, cash_back, round(paid_total - cash_back, 2),
        "示例退保：与其他保障重复", "购买前核对现有保单清单，避免保障重复",
    ])

    out_path = out_dir / "Insurance_Portfolio.xlsx"
    _save_workbook(wb, out_path, _fixed_xlsx_datetime(persona))
    return out_path, premium_by_month


# ---------------------------------------------------------------------------
# RSU (Excel) — RSU_transactions.xlsx
# ---------------------------------------------------------------------------

_RSU_TXN_HEADER = ["交易日期", "资产名称", "交易类型", "单位", "数量", "单位价格_USD", "总金额_USD", "手续费_USD", "备注"]


def emit_rsu(persona: dict[str, Any], out_dir: Path, rng: random.Random) -> tuple[Path, list]:
    asset_name = persona["rsu"]["asset_name"]
    sched = persona["rsu"]["vest_schedule"]
    snapshot_date: date = persona["timeline"]["snapshot_date"]
    price_lo, price_hi = persona["rsu"]["price_usd_range"]
    shares_per_vest = sched["shares_per_vest"]
    sell_to_cover = sched["sell_to_cover_shares"]

    vest_dates = []
    d = sched["start"]
    while d <= snapshot_date:
        vest_dates.append(d)
        d = _add_months(sched["start"], 3 * len(vest_dates))

    price = (price_lo + price_hi) / 2
    rows: list[tuple[datetime, str, str, str, float, float, float, Any, str]] = []
    vest_events: list[dict[str, Any]] = []
    for d in vest_dates:
        price = min(price_hi, max(price_lo, round(price + rng.uniform(-10, 10), 2)))
        vest_amount = round(shares_per_vest * price, 2)
        rows.append((
            _dt(d), asset_name, "RSU Vest", "Shares", float(shares_per_vest), price, vest_amount, None,
            f"RSU vest — {shares_per_vest} shares released",
        ))
        vest_events.append({
            "date": d, "shares": float(shares_per_vest), "net_shares": float(shares_per_vest - sell_to_cover),
            "price_usd": price, "amount_usd": vest_amount,
        })
        sell_amount = round(-sell_to_cover * price, 2)
        fee = round(rng.uniform(3, 8), 2)
        rows.append((
            _dt(d), asset_name, "Sell", "Shares", float(-sell_to_cover), price, sell_amount, fee,
            f"Sell-to-cover for vest dated {d.isoformat()}",
        ))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transactions"
    ws.append(_RSU_TXN_HEADER)
    for r in rows:
        ws.append(list(r))

    ws_n = wb.create_sheet("Notes")
    ws_n.append(["Column reference (synthetic demo data)"])
    ws_n.append(["交易日期: ISO date of the vest or sell execution."])
    ws_n.append(["交易类型: RSU Vest (shares released) or Sell (sell-to-cover)."])
    ws_n.append(["数量: positive for vests, negative for sell-to-cover legs."])
    ws_n.append(["单位价格_USD / 总金额_USD: USD, no currency conversion applied here."])

    out_path = out_dir / "RSU_transactions.xlsx"
    _save_workbook(wb, out_path, _fixed_xlsx_datetime(persona))
    return out_path, vest_events


# ---------------------------------------------------------------------------
# CN Fund (Excel) — funding_transactions.xlsx
# ---------------------------------------------------------------------------
#
# Real-file verification note (checked directly via openpyxl on
# tests/fixtures/readers/funding_transactions.xlsx before writing this
# emitter): raw_holdings_paste and raw_transactions_paste are exactly 8 and
# 10 columns wide respectively (ws.max_column) — 关联银行卡 is a single
# combined "BankName | cardtail" string column, there is no separate
# trailing unnamed card-tail column beyond it. The 基金交易记录 sheet's
# 基金代码 column is stored as TEXT with leading zeros already stripped
# (cell.data_type == 's', not 'n') — not a numeric-typed cell — but
# config/readers/cn_fund.yaml documents the engine's mainline behavior as
# reading that column WITHOUT a str converter, landing as int64. We
# reproduce that mainline path (native Excel NUMBER cells) rather than the
# fixture's specific text-with-stripped-zeros artifact — both exercise the
# same normalize_fund_code() zero-pad recovery, but the numeric-cell form
# is what the engine actually produces when no converter is applied.

_CN_FUND_HOLDINGS_RAW_HEADER = ["基金代码", "基金简称", "基金类型", "净值日期", "单位净值", "持有份额", "参考市值", "关联银行卡"]
_CN_FUND_TXN_RAW_HEADER = ["确认日期", "基金代码", "基金简称", "业务类型", "确认状态", "确认份额", "确认金额", "手续费", "确认净值", "关联银行卡"]
_CN_FUND_HOLDINGS_HEADER = ["Asset_ID", "Asset_Name", "Asset_Type_Raw", "Snapshot_Date", "Market_Price_Unit", "Quantity", "Market_Value_Raw"]
_CN_FUND_TXN_HEADER = ["交易日期", "基金代码", "基金名称", "操作类型", "交易金额", "交易份额", "交易时基金单位净值", "手续费", "交易原因"]

# Processed-sheet 操作类型 vocabulary (CN_FUND_TYPE_MAP_SEED keys) + one
# value deliberately absent from the map, to exercise the unmapped->'other'
# fallback in cn_fund_transactions_from_sheet.
_CN_FUND_UNKNOWN_TYPE = "定期定额扣款"
_CN_FUND_MEMO_POOL = [
    "定投扣款", "手动申购", "计划赎回", "组合再平衡", "转换份额调整", "季度分红发放",
]


def emit_cn_fund(persona: dict[str, Any], out_dir: Path, rng: random.Random) -> Path:
    catalog = persona["cn_funds"]["catalog"]
    linked_banks = persona["cn_funds"]["linked_banks"]
    txn_cfg = persona["cn_funds"]["transactions"]
    months = txn_cfg["months_of_history"]
    approx_rows = txn_cfg["approx_rows"]
    required_types = txn_cfg["required_types"]
    holdings_summary_rows = persona["cn_funds"]["holdings_summary_rows"]
    snapshot_date: date = persona["timeline"]["snapshot_date"]
    start_date = _add_months(snapshot_date, -months)

    # Base NAV per fund: money-market (货币型) funds sit near 1.0; others
    # get a plausible NAV in a wider band, walked per-transaction later.
    base_price: dict[str, float] = {}
    for f in catalog:
        if f["type"] == "货币型":
            base_price[f["code"]] = round(rng.uniform(0.98, 1.02), 4)
        else:
            base_price[f["code"]] = round(rng.uniform(0.9, 3.2), 4)

    # A "plausible subset" of fund x bank pairings — not the full cross
    # product. The first fund gets both banks; the rest alternate.
    pairings: list[tuple[dict, dict]] = [(catalog[0], linked_banks[0]), (catalog[0], linked_banks[1])]
    for i, f in enumerate(catalog[1:], start=1):
        pairings.append((f, linked_banks[i % len(linked_banks)]))

    # --- raw_holdings_paste ---
    raw_holdings_rows = []
    for fund, bank in pairings:
        price = round(base_price[fund["code"]] * rng.uniform(0.97, 1.03), 4)
        qty = round(rng.uniform(2000, 90000), 2)
        card = f"{bank['bank']} | {bank['card_tail']}"
        raw_holdings_rows.append([
            fund["code"], fund["name"], fund["type"], _dt(snapshot_date), price, qty,
            round(qty * price, 2), card,
        ])

    # --- raw_transactions_paste ---
    raw_txn_types = ["买基金", "买基金", "卖基金", "超级转换份额调减"]
    raw_txn_rows = []
    raw_txn_dates = _spread_dates(rng, snapshot_date - timedelta(days=45), snapshot_date - timedelta(days=1), 15)
    for d in raw_txn_dates:
        fund, bank = rng.choice(pairings)
        raw_type = rng.choice(raw_txn_types)
        price = round(base_price[fund["code"]] * rng.uniform(0.95, 1.05), 4)
        qty = round(rng.uniform(500, 60000), 2)
        amount = round(qty * price, 2)
        fees = round(rng.uniform(0, 20), 2) if rng.random() > 0.3 else 0
        card = f"{bank['bank']} | {bank['card_tail']}"
        raw_txn_rows.append([_dt(d), fund["code"], fund["name"], raw_type, "成功", qty, amount, fees, price, card])
    raw_txn_rows.sort(key=lambda r: r[0], reverse=True)

    # --- 基金交易记录 (processed transactions) — drives 基金持仓汇总 reconciliation ---
    type_pool = ["申购", "申购", "申购", "赎回", "赎回"]
    net_qty_by_code: dict[str, float] = {f["code"]: 0.0 for f in catalog}
    price_walk: dict[str, float] = dict(base_price)
    txn_rows = []

    rows_per_fund = max(4, approx_rows // len(catalog))
    for fund in catalog:
        code = fund["code"]
        dates = _spread_dates(rng, start_date, snapshot_date - timedelta(days=1), rows_per_fund)
        for d in dates:
            price_walk[code] = round(max(0.5, price_walk[code] + rng.uniform(-0.03, 0.03)), 4)
            raw_type = rng.choice(type_pool)
            qty = round(rng.uniform(500, 30000), 2)
            amount = round(qty * price_walk[code], 2)
            fees = round(rng.uniform(0, 25), 2) if rng.random() > 0.4 else 0
            memo = rng.choice(_CN_FUND_MEMO_POOL)
            txn_rows.append([_dt(d), code, fund["name"], raw_type, amount, qty, price_walk[code], fees, memo])
            if raw_type in ("申购", "买基金", "超级转换-转入", "超级转换份额调增", "活期宝即充即用"):
                net_qty_by_code[code] += qty
            elif raw_type in ("赎回", "卖基金", "超级转换-转出", "超级转换份额调减", "快速取现"):
                net_qty_by_code[code] -= qty

    # Explicitly inject every persona-required type (+ the unmapped one) at
    # least once, spread across funds so no single fund carries them all.
    inject_types = list(required_types) + [_CN_FUND_UNKNOWN_TYPE]
    for i, rtype in enumerate(inject_types):
        fund = catalog[i % len(catalog)]
        code = fund["code"]
        d = snapshot_date - timedelta(days=2 + i)
        price_walk[code] = round(max(0.5, price_walk[code] + rng.uniform(-0.02, 0.02)), 4)
        qty = round(rng.uniform(500, 15000), 2)
        amount = round(qty * price_walk[code], 2)
        fees = round(rng.uniform(0, 10), 2)
        memo = rng.choice(_CN_FUND_MEMO_POOL)
        txn_rows.append([_dt(d), code, fund["name"], rtype, amount, qty, price_walk[code], fees, memo])
        if rtype in ("申购", "买基金", "超级转换-转入", "超级转换份额调增", "活期宝即充即用"):
            net_qty_by_code[code] += qty
        elif rtype in ("赎回", "卖基金", "超级转换-转出", "超级转换份额调减", "快速取现"):
            net_qty_by_code[code] -= qty
        # 现金分红 / the unmapped filler type: share count unaffected.

    for rt in required_types:
        assert any(r[3] == rt for r in txn_rows), f"cn_fund required_type never emitted: {rt}"
    assert any(r[3] == _CN_FUND_UNKNOWN_TYPE for r in txn_rows), "cn_fund unmapped type never emitted"

    txn_rows.sort(key=lambda r: r[0])

    # --- 基金持仓汇总 (processed holdings) ---
    # Every fund's latest row must land within the QDII 2-day window of the
    # sheet's global max Snapshot_Date, or cn_fund_holdings_from_sheet drops
    # it. Give every fund a "latest" row on/near snapshot_date, plus a few
    # older history rows (outside the window, filtered out — realism only).
    holdings_rows = []
    for fund in catalog:
        code = fund["code"]
        qty = max(round(net_qty_by_code[code], 4), round(rng.uniform(1000, 5000), 4))
        price = price_walk[code]
        holdings_rows.append([
            code, fund["name"], fund["type"], _dt(snapshot_date), price, qty, round(qty * price, 2),
        ])
    n_history = max(0, holdings_summary_rows - len(catalog))
    history_funds = (catalog * ((n_history // len(catalog)) + 1))[:n_history]
    for i, fund in enumerate(history_funds):
        code = fund["code"]
        d = snapshot_date - timedelta(days=30 * (i % 5 + 1))
        price = round(base_price[code] * rng.uniform(0.9, 1.1), 4)
        qty = round(rng.uniform(1000, 20000), 2)
        holdings_rows.append([code, fund["name"], fund["type"], _dt(d), price, qty, round(qty * price, 2)])

    wb = openpyxl.Workbook()
    ws_rh = wb.active
    ws_rh.title = "raw_holdings_paste"
    ws_rh.append(_CN_FUND_HOLDINGS_RAW_HEADER)
    for r in raw_holdings_rows:
        ws_rh.append(r)

    ws_rt = wb.create_sheet("raw_transactions_paste")
    ws_rt.append(_CN_FUND_TXN_RAW_HEADER)
    for r in raw_txn_rows:
        ws_rt.append(r)

    ws_h = wb.create_sheet("基金持仓汇总")
    ws_h.append(_CN_FUND_HOLDINGS_HEADER)
    for r in holdings_rows:
        ws_h.append(r)

    ws_t = wb.create_sheet("基金交易记录")
    ws_t.append(_CN_FUND_TXN_HEADER)
    for r in txn_rows:
        # 基金代码 written as a native int cell — see module note above on
        # why this (not a stripped-zero text cell) is the faithful
        # reproduction of the engine's documented no-converter read path.
        code_int = int(r[1])
        ws_t.append([r[0], code_int, r[2], r[3], r[4], r[5], r[6], r[7], r[8]])

    out_path = out_dir / "funding_transactions.xlsx"
    _save_workbook(wb, out_path, _fixed_xlsx_datetime(persona))
    return out_path


# ---------------------------------------------------------------------------
# Financial Summary (Excel) — Financial_Summary_new.xlsx
# ---------------------------------------------------------------------------
#
# Real-file verification note (openpyxl on the real fixture, structure only):
#   - Both sheets: header=row4 (0-indexed 3); rows 1-3 carry merged group
#     labels above it; data starts row5.
#   - melt_financial_summary_holdings (src/sources/reader_hooks.py) does an
#     EXTRA `.iloc[3:]` trim AFTER header=3 — confirmed empirically by
#     running the real reader on the real fixture: the earliest melted
#     holdings date is 2020-02, three months after the real sheet's actual
#     first data row (2019-11). This silently drops the first 3 rows of
#     EVERY 资产负债 melt, unconditionally — not something to "fix" here
#     (out of WS-1 scope), but it means our own 24-month window's first 3
#     months (2024-07..2024-09) will not appear in melted holdings, and the
#     same trim exists in financial_summary_sync._read_fs_sheet for 月度收支.
#     Any month-spot-check in the verification gate must use a later month.
#   - No `fs_column`-equivalent seed exists for 资产负债's 合计 columns (only
#     IE_COLUMN_SEED exists, for 月度收支). The 合计 formulas below were
#     reverse-engineered from one real data row (values only, structure is
#     public/generic accounting math) and cross-checked to reproduce exactly
#     except one artifact: real 合计流动资产 silently includes the
#     投资资产_黄金_纸黄金(克) GRAM QUANTITY column (a stray range in the
#     owner's own SUM formula) — persona.invariants #1 wants a clean
#     leaf-sum, so that artifact is deliberately NOT reproduced here.

# Row1-3 merged-group-label arrays, index-aligned with the persona column
# lists (WS-0 confirmed persona column ORDER is identical to the real
# workbook's, only names differ on the flagged renames) — generic
# accounting category labels (资产/负债/活期/股票基金/…), not owner data.
_FS_BS_ROW1 = [None, "资产", None, None, None, None, None, None, None, None, None, None, None, None, None, None,
               None, None, None, None, None, None, None, None, None, "负债", None, None, None, None, None, None,
               None, None, None, "资产情况", None, None, None, None, None, None, None, None, None, None]
_FS_BS_ROW2 = [None, "RMB现金资产", None, None, None, None, "美元账户", None, None, None, None, "投资资产", None,
               None, None, None, None, None, None, None, None, None, None, None, "固定资产", "短期负债", None,
               None, None, None, None, None, "长期负债", None, "其他", None, None, None, None, None, None, None,
               None, None, None, None]
_FS_BS_ROW3 = [None, "活期", None, None, None, None, None, None, None, None, None, "股份投资", "股票基金",
               "股票基金", None, "RSU", None, "银行理财", "养老金", "黄金", None, None, "商业保险", "另类投资",
               "房产", "信用卡", None, None, None, None, None, None, "贷款", None, None, None, None, None, None,
               None, None, None, None, None, None, None]

_FS_IE_ROW1 = [None, "收入", None, None, None, None, None, None, None, None, None, None, None, None, None,
               "支出", None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
               None, None, None, None, None, None, None, None, None, None, None, None, None]
_FS_IE_ROW2 = [None, "主动收入", None, None, None, None, None, None, "被动收入", None, None, None, "合计", None,
               None, "必要开支", None, None, None, None, None, None, None, "非必要开支", None, None, None, None,
               None, "投资理财", None, None, None, None, None, "合计", None, None, None, None, None, None, None]
_FS_IE_ROW3 = [None, "收入_主动收入_工资", "USD", "收入_主动收入_RSU", "收入_主动收入_报销", "收入_主动收入_福利",
               "收入_主动收入_公积金", "收入_主动收入_其他偶然", "收入_被动收入_银行理财", None, None,
               "收入_被动收入_", "主动收入", "被动收入", "总收入", "日常支出", None, None, "保险", None, None,
               "贷款", "其他", "非日常支出", None, None, None, None, "工作支出", "银行理财", "黄金", None,
               "股票基金", None, None, "必要支出", "非必要支出", "工作支出", "理财", "总支出", None, None, None]


def _linear_walk(rng: random.Random, start: float, end: float, n: int,
                  jitter_frac: float = 0.06, floor: float = 0.0) -> list[float]:
    """n values interpolating start->end with proportional jitter, floored >=0."""
    if n <= 1:
        return [round(max(floor, end), 2)]
    out = []
    for i in range(n):
        base = start + (end - start) * i / (n - 1)
        jitter = base * jitter_frac * rng.uniform(-1, 1)
        out.append(round(max(floor, base + jitter), 2))
    return out


def _month_grid(first_month: date, n: int) -> list[date]:
    return [_add_months(first_month, i) for i in range(n)]


def emit_financial_summary(
    persona: dict[str, Any], out_dir: Path, rng: random.Random, *,
    gold_txns: list[dict[str, Any]], premium_by_month: dict[date, dict[str, float]],
    vest_events: list[dict[str, Any]], wire_transfers: list[dict[str, Any]],
) -> Path:
    fs = persona["financial_summary"]
    bs_cfg = fs["balance_sheet"]
    ie_cfg = fs["income_expense"]
    bs_cols = bs_cfg["columns"]
    ie_cols = ie_cfg["columns"]
    months = bs_cfg["months"]
    assert months == ie_cfg["months"], "balance_sheet/income_expense month counts must match"

    start_month_str: str = persona["timeline"]["start_month"]
    y, m = (int(x) for x in start_month_str.split("-"))
    first_month = date(y, m, 1)
    grid = _month_grid(first_month, months)  # 24 date objects, first-of-month

    usd_cny_base = persona["timeline"]["usd_cny"]
    gold_price_base = persona["timeline"]["gold_cny_per_gram"]
    price_lo, price_hi = persona["rsu"]["price_usd_range"]
    targets = persona["targets"]

    # ---- Shared reference series (identical across both sheets) ----------
    fx_rate = [usd_cny_base]
    for _ in range(1, months):
        fx_rate.append(round(min(7.4, max(6.8, fx_rate[-1] + rng.uniform(-0.03, 0.03))), 4))

    gold_price = [gold_price_base]
    for _ in range(1, months):
        gold_price.append(round(min(950, max(600, gold_price[-1] + rng.uniform(-15, 15))), 2))

    stock_price = [(price_lo + price_hi) / 2] * months
    walk = stock_price[0]
    for i in range(months):
        walk = min(price_hi, max(price_lo, round(walk + rng.uniform(-8, 8), 2)))
        stock_price[i] = walk

    # ---- Cross-file reconciled monthly series -----------------------------
    vest_by_month: dict[date, dict[str, float]] = {}
    for v in vest_events:
        ym = date(v["date"].year, v["date"].month, 1)
        if ym in grid:
            vest_by_month[ym] = v
            stock_price[grid.index(ym)] = v["price_usd"]  # glue the reference series to the actual vest price

    wire_by_month: dict[date, float] = {}
    for w in wire_transfers:
        ym = date(w["date"].year, w["date"].month, 1)
        if ym in grid:
            wire_by_month[ym] = wire_by_month.get(ym, 0.0) + w["amount_usd"]

    cmb_gold_buy_by_month: dict[date, float] = {}
    gold_grams_cumulative: list[float] = []
    running_grams = 0.0
    sorted_gold_txns = sorted(gold_txns, key=lambda t: t["date"])
    for ym in grid:
        month_end = _add_months(ym, 1)
        while sorted_gold_txns and sorted_gold_txns[0]["date"] < month_end:
            t = sorted_gold_txns.pop(0)
            sign = 1.0 if t["type"] in ("买入", "结息") else -1.0
            running_grams += sign * t["qty"]
            if t["account"] == "招行" and t["type"] == "买入":
                cmb_gold_buy_by_month[ym] = cmb_gold_buy_by_month.get(ym, 0.0) + t["amount"]
        gold_grams_cumulative.append(round(max(0.0, running_grams), 4))

    # RSU cumulative net-share holding (vest - sell_to_cover), for the FS's
    # own informational RSU balance column (not invariant-required, but
    # cheap realism using data we already have).
    rsu_net_shares_cumulative: list[float] = []
    running_shares = 0.0
    for ym in grid:
        if ym in vest_by_month:
            running_shares += vest_by_month[ym]["net_shares"]
        rsu_net_shares_cumulative.append(running_shares)

    # ---- 资产负债 (balance sheet) ------------------------------------------
    mortgage_start = bs_cfg["mortgage_start_cny"]
    mortgage_paydown = round(mortgage_start * 0.12 / months, 2)  # fixed, strictly monotonic
    property_value = bs_cfg["property_value_cny"]
    pension_monthly = bs_cfg["pension_monthly_cny"]

    bs: dict[str, list[float]] = {}
    bs["日期"] = grid
    bs["RMB现金现金"] = _linear_walk(rng, 1000, 3000, months, floor=500)
    bs["RMB存款_中行"] = _linear_walk(rng, 15000, 40000, months, floor=1000)
    bs["RMB存款_招行"] = _linear_walk(rng, 12000, 35000, months, floor=1000)
    bs["RMB存款_北京银行"] = _linear_walk(rng, 3000, 8000, months, floor=200)
    usd_boc = _linear_walk(rng, 1800, 2800, months, floor=100)
    usd_chase = _linear_walk(rng, 5500, 9000, months, floor=100)
    usd_discover = _linear_walk(rng, 1500, 2600, months, floor=100)
    usd_fund_usd = _linear_walk(rng, 12000, 19000, months, floor=500)
    usd_cc_annual = _linear_walk(rng, 60, 110, months, floor=0)
    usd_cc_bill = _linear_walk(rng, 300, 1200, months, jitter_frac=0.3, floor=0)
    bs["美元存款_中行_USD"] = usd_boc
    bs["美元存款_中行"] = [round(usd_boc[i] * fx_rate[i], 2) for i in range(months)]
    bs["美元存款_Chase_USD"] = usd_chase
    bs["美元存款_Chase"] = [round(usd_chase[i] * fx_rate[i], 2) for i in range(months)]
    bs["美元存款_Discover/Citi_USD"] = usd_discover
    bs["美元存款_Discover/Citi"] = [round(usd_discover[i] * fx_rate[i], 2) for i in range(months)]
    bs["创业股权投资"] = [None] * months  # unmapped legacy column; real fixture is empty too
    bs["投资资产_股票基金_A股基金"] = _linear_walk(rng, 600000, 1300000, months, floor=0)
    bs["投资资产_股票基金_美股基金_USD"] = usd_fund_usd
    bs["投资资产_股票基金_美股基金"] = [round(usd_fund_usd[i] * fx_rate[i], 2) for i in range(months)]
    bs["投资资产_公司RSU_Amazon Stock_USD"] = [round(rsu_net_shares_cumulative[i] * stock_price[i], 2) for i in range(months)]
    bs["投资资产_公司RSU_Amazon Stock"] = [round(bs["投资资产_公司RSU_Amazon Stock_USD"][i] * fx_rate[i], 2) for i in range(months)]
    bs["投资资产_银行理财_招行"] = _linear_walk(rng, 80000, 180000, months, floor=0)
    bs["投资资产_存款基金_个人养老金"] = [round(pension_monthly * (i + 1), 2) for i in range(months)]
    bs["投资资产_黄金_纸黄金(元)"] = [round(gold_grams_cumulative[i] * gold_price[i], 2) for i in range(months)]
    bs["投资资产_黄金_纸黄金(克)"] = gold_grams_cumulative
    bs["投资资产_黄金_黄金ETF"] = [None] * months  # persona has no gold ETF holding
    bs["投资资产_长期保险_安泰人生"] = _linear_walk(rng, 15000, 22000, months, floor=0)
    bs["投资资产_另类投资_公司股份投资"] = [None] * months  # owner-reviewed, no reader coverage (matches real)
    bs["固定资产_房产_阳光花园"] = [property_value] * months
    bs["短期负债_信用卡_招行"] = _linear_walk(rng, 1000, 6000, months, jitter_frac=0.4, floor=0)
    bs["短期负债_信用卡_中行"] = _linear_walk(rng, 1000, 5000, months, jitter_frac=0.4, floor=0)
    bs["短期负债_信用卡_中信"] = _linear_walk(rng, 2000, 9000, months, jitter_frac=0.4, floor=0)
    bs["短期负债_信用卡_美国信用卡年费_USD"] = usd_cc_annual
    bs["短期负债_信用卡_美国信用卡年费"] = [round(usd_cc_annual[i] * fx_rate[i], 2) for i in range(months)]
    bs["短期负债_信用卡_美国信用卡账单_USD"] = usd_cc_bill
    bs["短期负债_信用卡_美国信用卡账单"] = [round(usd_cc_bill[i] * fx_rate[i], 2) for i in range(months)]
    bs["长期负债_房贷"] = [round(mortgage_start - mortgage_paydown * i, 2) for i in range(months)]
    bs["长期负债_车贷"] = [None] * months
    bs["其他负债"] = [None] * months
    bs["Amazon Stock"] = stock_price
    bs["USD Rate"] = fx_rate

    # ---- Calibrate the FREE (non-reconciled) asset columns to hit the
    # persona net-worth target within +/-5% at the final month, WITHOUT
    # disturbing any cross-file-reconciled column above. ---------------------
    free_asset_cols = [
        "RMB现金现金", "RMB存款_中行", "RMB存款_招行", "RMB存款_北京银行",
        "投资资产_股票基金_A股基金", "投资资产_银行理财_招行", "投资资产_长期保险_安泰人生",
    ]

    def _bs_totals(i: int) -> dict[str, float]:
        liquid = sum(
            (bs[c][i] or 0.0) for c in (
                "RMB现金现金", "RMB存款_中行", "RMB存款_招行", "RMB存款_北京银行",
                "美元存款_中行", "美元存款_Chase", "美元存款_Discover/Citi",
                "投资资产_股票基金_A股基金", "投资资产_股票基金_美股基金",
                "投资资产_公司RSU_Amazon Stock", "投资资产_银行理财_招行",
                "投资资产_存款基金_个人养老金", "投资资产_黄金_纸黄金(元)",
                "投资资产_长期保险_安泰人生",
            )
        )
        short_term_liab = sum(
            (bs[c][i] or 0.0) for c in (
                "短期负债_信用卡_招行", "短期负债_信用卡_中行", "短期负债_信用卡_中信",
                "短期负债_信用卡_美国信用卡年费", "短期负债_信用卡_美国信用卡账单",
            )
        )
        total_liab = short_term_liab + sum(
            (bs[c][i] or 0.0) for c in ("长期负债_房贷", "长期负债_车贷", "其他负债")
        )
        total_assets = liquid + bs["固定资产_房产_阳光花园"][i]
        net_worth = total_assets - total_liab
        return {
            "liquid": liquid, "short_term_liab": short_term_liab, "total_liab": total_liab,
            "total_assets": total_assets, "net_worth": net_worth,
        }

    last = months - 1
    raw_totals_final = _bs_totals(last)
    raw_free_final = sum(bs[c][last] for c in free_asset_cols)
    target_net_worth = targets["net_worth_cny"]
    scale = 1.0 + (target_net_worth - raw_totals_final["net_worth"]) / raw_free_final
    assert scale > 0, f"cn_fund calibration scale went non-positive ({scale}); adjust base magnitudes"
    for c in free_asset_cols:
        bs[c] = [round(v * scale, 2) for v in bs[c]]

    # ---- 资产负债 computed/ratio columns (invariant #1: leaf-sum, exact) ---
    bs["合计流动资产"] = [round(_bs_totals(i)["liquid"], 2) for i in range(months)]
    bs["合计负债资产"] = [round(_bs_totals(i)["total_liab"], 2) for i in range(months)]
    bs["合计流动净资产"] = [round(bs["合计流动资产"][i] - _bs_totals(i)["short_term_liab"], 2) for i in range(months)]
    bs["合计总资产"] = [round(bs["合计流动资产"][i] + bs["固定资产_房产_阳光花园"][i], 2) for i in range(months)]
    bs["合计净资产"] = [round(bs["合计总资产"][i] - bs["合计负债资产"][i], 2) for i in range(months)]
    bs["资产负债率 50%"] = [
        round(bs["合计负债资产"][i] / bs["合计总资产"][i], 6) if bs["合计总资产"][i] else 0.0
        for i in range(months)
    ]
    bs["即付比例 70%"] = [
        round(bs["合计流动资产"][i] / _bs_totals(i)["short_term_liab"], 6)
        if _bs_totals(i)["short_term_liab"] else 0.0
        for i in range(months)
    ]
    invested_like = (
        "投资资产_股票基金_A股基金", "投资资产_股票基金_美股基金", "投资资产_公司RSU_Amazon Stock",
        "投资资产_银行理财_招行", "投资资产_黄金_纸黄金(元)", "投资资产_长期保险_安泰人生",
    )
    bs["投资比例 50%"] = [
        round(sum((bs[c][i] or 0.0) for c in invested_like) / bs["合计总资产"][i], 6)
        if bs["合计总资产"][i] else 0.0
        for i in range(months)
    ]

    # ---- Invariant self-checks (资产负债) -----------------------------------
    for i in range(months):
        assert abs(bs["合计流动净资产"][i] - (bs["合计流动资产"][i] - _bs_totals(i)["short_term_liab"])) < 0.01
        assert abs(bs["合计总资产"][i] - (bs["合计流动资产"][i] + property_value)) < 0.01
        assert abs(bs["合计净资产"][i] - (bs["合计总资产"][i] - bs["合计负债资产"][i])) < 0.01
    assert bs["固定资产_房产_阳光花园"] == [property_value] * months, "property value must stay constant"
    pension = bs["投资资产_存款基金_个人养老金"]
    assert all(pension[i] - pension[i - 1] == pension_monthly for i in range(1, months)), \
        "pension must grow by exactly pension_monthly_cny each month"
    mortgage = bs["长期负债_房贷"]
    assert all(mortgage[i] < mortgage[i - 1] for i in range(1, months)), "mortgage must decline monotonically"
    final_net_worth = bs["合计净资产"][last]
    assert abs(final_net_worth - target_net_worth) <= 0.05 * target_net_worth, (
        f"final net worth {final_net_worth} not within +/-5% of target {target_net_worth}"
    )
    for c in bs:
        if c in ("日期",):
            continue
        assert all(v is None or v >= -0.01 for v in bs[c]), f"{c} went negative"

    # ---- 月度收支 (income/expense) -----------------------------------------
    ie: dict[str, list[float]] = {}
    ie["日期"] = grid
    salary_base = targets["monthly_salary_cny"]
    ie["收入_主动收入_工资"] = _linear_walk(rng, salary_base * 0.97, salary_base * 1.03, months, jitter_frac=0.02, floor=0)

    rsu_usd = [round(vest_by_month[ym]["amount_usd"], 2) if ym in vest_by_month else 0.0 for ym in grid]
    ie["收入_主动收入_RSU_USD"] = rsu_usd
    ie["收入_主动收入_RSU"] = [round(rsu_usd[i] * fx_rate[i], 2) for i in range(months)]

    # Built as a single grid-ordered pass (not via an intermediate set) —
    # date.__hash__ is PYTHONHASHSEED-salted, so iterating a set of dates
    # (rather than just membership-testing it) would consume rng draws in a
    # different order every process, breaking cross-run determinism.
    reimb_amount: dict[date, float] = {}
    for ym in grid:
        if rng.random() < 0.3:
            reimb_amount[ym] = round(rng.uniform(500, 3000), 2)
    ie["收入_主动收入_报销"] = [reimb_amount.get(ym, 0.0) for ym in grid]
    ie["工作开支_出差/团建（全额报销）"] = [reimb_amount.get(ym, 0.0) for ym in grid]

    welfare_months = set(rng.sample(grid, k=max(1, months // 12)))
    ie["收入_主动收入_福利"] = [round(rng.uniform(60000, 150000), 2) if ym in welfare_months else 0.0 for ym in grid]
    fund_months = set(rng.sample(grid, k=max(1, months // 12)))
    ie["收入_主动收入_公积金"] = [round(rng.uniform(50000, 90000), 2) if ym in fund_months else 0.0 for ym in grid]
    bonus_months = set(rng.sample(grid, k=max(1, months // 24)))
    ie["收入_主动收入_其他偶然"] = [round(rng.uniform(5000, 30000), 2) if ym in bonus_months else 0.0 for ym in grid]

    ie["收入_被动收入_银行理财"] = [round(rng.uniform(0, 3000), 2) if rng.random() < 0.25 else 0.0 for _ in grid]
    gold_sell_amounts = []
    gold_sell_grams = []
    for i in grid:
        if rng.random() < 0.1:
            g = round(rng.uniform(3, 10), 4)
            gold_sell_grams.append(g)
            gold_sell_amounts.append(round(g * gold_price[grid.index(i)], 2))
        else:
            gold_sell_grams.append(0.0)
            gold_sell_amounts.append(0.0)
    ie["收入_被动收入_黄金卖出"] = gold_sell_amounts
    ie["收入_被动收入_黄金卖出(克)"] = gold_sell_grams
    ie["收入_被动收入_基金赎回"] = [round(rng.uniform(0, 5000), 2) if rng.random() < 0.15 else 0.0 for _ in grid]

    active_income_leaves = (
        "收入_主动收入_工资", "收入_主动收入_RSU", "收入_主动收入_报销", "收入_主动收入_福利",
        "收入_主动收入_公积金", "收入_主动收入_其他偶然",
    )
    passive_income_leaves = ("收入_被动收入_银行理财", "收入_被动收入_黄金卖出", "收入_被动收入_基金赎回")
    ie["主动收入合计"] = [round(sum(ie[c][i] for c in active_income_leaves), 2) for i in range(months)]
    ie["被动收入合计"] = [round(sum(ie[c][i] for c in passive_income_leaves), 2) for i in range(months)]
    ie["总收入合计"] = [round(ie["主动收入合计"][i] + ie["被动收入合计"][i], 2) for i in range(months)]

    essential_target = [round(rng.uniform(*targets["monthly_essential_spend_cny"]), 2) for _ in grid]
    # persona.insurance.premium_record_columns ARE the bare FS column suffixes
    # (安泰人生/公司团险/互联网保险) — the actual FS column names carry the
    # 必要开支_保险_ prefix; premium_by_month is keyed by the bare names.
    insurance_leaves_cols = list(persona["insurance"]["premium_record_columns"])
    insurance_cny_by_month = {
        ym: sum(premium_by_month.get(ym, {}).values()) for ym in grid
    }
    for col in insurance_leaves_cols:
        full_col = f"必要开支_保险_{col}"
        ie[full_col] = [round(premium_by_month.get(ym, {}).get(col, 0.0), 2) for ym in grid]

    mortgage_payment = [round(mortgage_paydown, 2)] * months
    ie["必要开支_贷款_房贷"] = mortgage_payment
    remaining_daily = [
        round(max(600.0, essential_target[i] - insurance_cny_by_month[grid[i]] - mortgage_payment[i]), 2)
        for i in range(months)
    ]
    ie["必要开支_日常支出_餐饮娱乐"] = [round(v * 0.40, 2) for v in remaining_daily]
    ie["必要开支_日常支出_房租水电"] = [round(v * 0.35, 2) for v in remaining_daily]
    ie["必要开支_日常支出_交通"] = [round(v * 0.15, 2) for v in remaining_daily]
    ie["必要开支_家庭及临时支出"] = [round(v * 0.10, 2) for v in remaining_daily]

    discretionary_target = [round(rng.uniform(*targets["monthly_discretionary_spend_cny"]), 2) for _ in grid]
    ie["非必要开支_旅行出游"] = [round(v * 0.35, 2) for v in discretionary_target]
    ie["非必要开支_护肤衣物"] = [round(v * 0.20, 2) for v in discretionary_target]
    ie["非必要开支_电子产品"] = [round(v * 0.20, 2) for v in discretionary_target]
    ie["非必要开支_运动健身健康"] = [round(v * 0.15, 2) for v in discretionary_target]
    ie["非必要开支_其他/娱乐"] = [round(v * 0.10, 2) for v in discretionary_target]

    ie["投资理财_银行理财_招行"] = [round(rng.uniform(0, 6000), 2) if rng.random() < 0.4 else 0.0 for _ in grid]
    ie["投资理财_黄金_招行纸黄金"] = [round(cmb_gold_buy_by_month.get(ym, 0.0), 2) for ym in grid]
    ie["投资理财_黄金_黄金ETF"] = [0.0] * months  # no ETF gold activity in persona
    schawab_usd = [round(wire_by_month.get(ym, 0.0), 2) for ym in grid]
    ie["投资理财_股票基金_Schawab_USD"] = schawab_usd
    ie["投资理财_股票基金_Schawab"] = [round(schawab_usd[i] * fx_rate[i], 2) for i in range(months)]
    ie["投资理财_股票基金_天天基金"] = [round(rng.uniform(0, 8000), 2) if rng.random() < 0.35 else 0.0 for _ in grid]

    essential_leaves = (
        "必要开支_日常支出_餐饮娱乐", "必要开支_日常支出_房租水电", "必要开支_日常支出_交通",
        f"必要开支_保险_{insurance_leaves_cols[0]}", f"必要开支_保险_{insurance_leaves_cols[1]}",
        f"必要开支_保险_{insurance_leaves_cols[2]}", "必要开支_贷款_房贷", "必要开支_家庭及临时支出",
    )
    discretionary_leaves = (
        "非必要开支_旅行出游", "非必要开支_护肤衣物", "非必要开支_电子产品",
        "非必要开支_运动健身健康", "非必要开支_其他/娱乐",
    )
    work_leaves = ("工作开支_出差/团建（全额报销）",)
    invested_leaves = (
        "投资理财_银行理财_招行", "投资理财_黄金_招行纸黄金", "投资理财_黄金_黄金ETF",
        "投资理财_股票基金_Schawab", "投资理财_股票基金_天天基金",
    )
    ie["必要支出"] = [round(sum(ie[c][i] for c in essential_leaves), 2) for i in range(months)]
    ie["非必要支出"] = [round(sum(ie[c][i] for c in discretionary_leaves), 2) for i in range(months)]
    ie["工作支出"] = [round(sum(ie[c][i] for c in work_leaves), 2) for i in range(months)]
    ie["理财"] = [round(sum(ie[c][i] for c in invested_leaves), 2) for i in range(months)]
    ie["总支出"] = [
        round(ie["必要支出"][i] + ie["非必要支出"][i] + ie["工作支出"][i] + ie["理财"][i], 2)
        for i in range(months)
    ]

    ie["参考_黄金价格_克价"] = gold_price
    ie["参考_Amazon Stock Price"] = stock_price
    ie["参考_美元汇率"] = fx_rate

    # ---- Invariant self-checks (月度收支) -----------------------------------
    for i in range(months):
        assert abs(ie["主动收入合计"][i] - sum(ie[c][i] for c in active_income_leaves)) < 0.01
        assert abs(ie["被动收入合计"][i] - sum(ie[c][i] for c in passive_income_leaves)) < 0.01
        assert abs(ie["总收入合计"][i] - (ie["主动收入合计"][i] + ie["被动收入合计"][i])) < 0.01
        assert abs(ie["必要支出"][i] - sum(ie[c][i] for c in essential_leaves)) < 0.01
        assert abs(ie["非必要支出"][i] - sum(ie[c][i] for c in discretionary_leaves)) < 0.01
        assert abs(ie["工作支出"][i] - sum(ie[c][i] for c in work_leaves)) < 0.01
        assert abs(ie["理财"][i] - sum(ie[c][i] for c in invested_leaves)) < 0.01
        assert abs(
            ie["总支出"][i] - (ie["必要支出"][i] + ie["非必要支出"][i] + ie["工作支出"][i] + ie["理财"][i])
        ) < 0.01, "总支出 must include 理财 (uis-never-reads-excel-aggregates)"
    for i, ym in enumerate(grid):
        if ym in vest_by_month:
            assert ie["收入_主动收入_RSU_USD"][i] > 0, f"RSU vest month {ym} must be nonzero"
        else:
            assert ie["收入_主动收入_RSU_USD"][i] == 0.0, f"non-vest month {ym} must be exactly zero"
        if ym in wire_by_month:
            assert ie["投资理财_股票基金_Schawab_USD"][i] > 0, f"wire month {ym} must be nonzero"
        else:
            assert ie["投资理财_股票基金_Schawab_USD"][i] == 0.0, f"non-wire month {ym} must be exactly zero"
    for i in range(months):
        assert abs(bs["投资资产_黄金_纸黄金(元)"][i] - bs["投资资产_黄金_纸黄金(克)"][i] * gold_price[i]) < 0.01

    # ---- Assemble columns in persona's declared order, write workbook -----
    wb = openpyxl.Workbook()
    ws_bs = wb.active
    ws_bs.title = "资产负债"
    ws_bs.append(_FS_BS_ROW1)
    ws_bs.append(_FS_BS_ROW2)
    ws_bs.append(_FS_BS_ROW3)
    ws_bs.append(bs_cols)
    for i in range(months):
        row = []
        for col in bs_cols:
            if col == "日期":
                row.append(_dt(grid[i]))
            else:
                v = bs.get(col, [None] * months)[i]
                row.append(v)
        ws_bs.append(row)

    ws_ie = wb.create_sheet("月度收支")
    ws_ie.append(_FS_IE_ROW1)
    ws_ie.append(_FS_IE_ROW2)
    ws_ie.append(_FS_IE_ROW3)
    ws_ie.append(ie_cols)
    for i in range(months):
        row = []
        for col in ie_cols:
            key = col.strip()  # persona keeps 参考_美元汇率 without the real file's trailing space
            if col == "日期":
                row.append(_dt(grid[i]))
            else:
                v = ie.get(key, [None] * months)[i]
                row.append(v)
        ws_ie.append(row)

    # ---- Dummy year-summary sheets (not consumed by anything) -------------
    for sheet_name in fs["dummy_year_sheets"]:
        ws_y = wb.create_sheet(sheet_name)
        ws_y.append(["月份", "备注"])
        ws_y.append(["示例年度汇总", "占位表，未被任何读取逻辑使用"])
        ws_y.append(["用途", "保持工作簿结构完整性"])

    out_path = out_dir / "Financial_Summary_new.xlsx"
    _save_workbook(wb, out_path, _fixed_xlsx_datetime(persona))
    return out_path


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------


def generate_all(persona: dict[str, Any], out_dir: Path) -> list[Path]:
    if out_dir.exists():
        shutil.rmtree(out_dir)
    out_dir.mkdir(parents=True)

    rng = random.Random(persona["rng_seed"])

    paths: list[Path] = []
    paths.extend(emit_ibkr(persona, out_dir, rng))
    paths.append(emit_schwab_positions(persona, out_dir, rng))
    schwab_txn_path, wire_transfers = emit_schwab_transactions(persona, out_dir, rng)
    paths.append(schwab_txn_path)
    gold_path, gold_txns = emit_gold(persona, out_dir, rng)
    paths.append(gold_path)
    insurance_path, premium_by_month = emit_insurance(persona, out_dir, rng)
    paths.append(insurance_path)
    rsu_path, vest_events = emit_rsu(persona, out_dir, rng)
    paths.append(rsu_path)
    paths.append(emit_cn_fund(persona, out_dir, rng))
    paths.append(emit_financial_summary(
        persona, out_dir, rng,
        gold_txns=gold_txns, premium_by_month=premium_by_month,
        vest_events=vest_events, wire_transfers=wire_transfers,
    ))
    return paths


def install_fixtures(out_dir: Path, sources: list[str], fixture_dir: Path = FIXTURE_DIR) -> list[Path]:
    """Copy generated files for the given sources into tests/fixtures/readers/,
    at their exact real filenames (creating ibkr/ibkr_trades subdirs as needed).

    Reproducible swap step for WS-1.5+ — run generate.py, then install just the
    sources being swapped this phase. Does not delete anything (e.g. a stale
    differently-named schwab fixture from before an account-tail rename is
    left for the operator to `git rm` explicitly).
    """
    unknown = set(sources) - set(INSTALL_GLOBS)
    if unknown:
        raise SystemExit(f"--install: unknown source(s) {sorted(unknown)}; valid: {sorted(INSTALL_GLOBS)}")

    installed: list[Path] = []
    for source in sources:
        for pattern in INSTALL_GLOBS[source]:
            matches = sorted(out_dir.glob(pattern))
            if not matches:
                raise SystemExit(f"--install: no file matched '{pattern}' for source '{source}' under {out_dir}")
            for src_path in matches:
                rel = src_path.relative_to(out_dir)
                dest = fixture_dir / rel
                dest.parent.mkdir(parents=True, exist_ok=True)
                shutil.copyfile(src_path, dest)
                installed.append(dest)
    return installed


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--out-dir", type=Path, default=DEFAULT_OUT_DIR)
    parser.add_argument("--persona", type=Path, default=PERSONA_PATH)
    parser.add_argument(
        "--install", action="store_true",
        help="After generating, copy the selected --sources' files into tests/fixtures/readers/.",
    )
    parser.add_argument(
        "--sources", default="",
        help=f"Comma-separated sources for --install. One of: {','.join(INSTALL_GLOBS)}.",
    )
    args = parser.parse_args()

    persona = load_persona(args.persona)
    paths = generate_all(persona, args.out_dir)
    for p in paths:
        print(f"wrote {p.relative_to(args.out_dir.parent)}")

    if args.install:
        sources = [s.strip() for s in args.sources.split(",") if s.strip()]
        if not sources:
            raise SystemExit("--install requires --sources (comma-separated; see --help)")
        installed = install_fixtures(args.out_dir, sources)
        for p in installed:
            print(f"installed {p.relative_to(FIXTURE_DIR.parent.parent.parent)}")


if __name__ == "__main__":
    main()
